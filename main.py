from kivymatplotlib.backend_kivyagg import FigureCanvasKivyAgg
from matplotlib import pyplot as pl
import matplotlib
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as getcl
from openpyxl.styles import Alignment,Font,Border,Side,PatternFill
import kivy
import kivymd
from kivy.utils import platform
from kivy.uix.bubble import Bubble,BubbleButton
from kivy.uix.label import Label
from kivy.uix.gridlayout import GridLayout
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.image import Image
from kivy.metrics import dp,sp
from kivy.uix.scrollview import ScrollView
from kivy.uix.carousel import Carousel
from kivy.core.window import Window
from kivy.uix.screenmanager import ScreenManager,Screen
from kivymd.uix.datatables import MDDataTable as DT
from kivy.graphics import RoundedRectangle,Color,Rectangle
from kivymd.app import MDApp
from kivymd.toast import toast as thongbao
from kivymd.uix.list import OneLineIconListItem,IconLeftWidget
from kivymd.uix.filemanager import MDFileManager
from kivymd.uix.dialog import MDDialog
from kivymd.uix.label import MDLabel
from kivymd.uix.tab import MDTabsBase,MDTabs
from kivymd.uix.button import MDFlatButton, MDRectangleFlatButton,MDIconButton,MDFillRoundFlatButton,MDFloatingActionButtonSpeedDial
from kivymd.uix.toolbar import MDToolbar,MDBottomAppBar
from kivymd.uix.textfield import MDTextField,MDTextFieldRound
from kivymd.uix.navigationdrawer import MDNavigationDrawer
from kivymd.uix.menu import MDDropdownMenu
from kivymd.uix.floatlayout import FloatLayout
#Class-----------------------------------------------------
class Loihs:
	def __init__(self,TenLoi,date):
		self.ten=TenLoi
		self.date=date
class Loi:
	def __init__(self,TenLoi,Diem):
		self.ten=TenLoi
		self.diem=Diem
class Cotdiemhs:
	def __init__(self,TenCotDiem):
		self.diem=[]
		self.ten=TenCotDiem
	def tbcotdiem(self,hocki,tenmon):
		heso=float(app.gethesomh(hocki,tenmon,self.ten))
		tong=0
		tongheso=0
		for diem in self.diem:
			tong+=float(diem)
			tongheso+=heso
		try:
			return round(tong/tongheso,1)
		except ZeroDivisionError:
			return 0
class Diemhs:
	def __init__(self,TenMon):
		self.ten=TenMon
		self.dhk1=[]
		self.dhk2=[]
	def dtbmon(self,hocki):
		tbmhk=0
		tongheso=0
		if hocki == 1:
			for cotdiem in self.dhk1:
				hesoo=app.gethesomh(1,self.ten,cotdiem.ten)
				for diemso in cotdiem.diem:
					if diemso=="":diemso=0
					tbmhk += float(diemso)*hesoo
					tongheso+=hesoo
		else:
			for cotdiem in self.dhk2:
				hesoo=app.gethesomh(2,self.ten,cotdiem.ten)
				for diemso in cotdiem.diem:
					if diemso=="":diemso=0
					tbmhk += float(diemso)*hesoo
					tongheso+=hesoo
		for mon in app.monhoc:
			if mon.ten == self.ten:
				if hocki==1:
					for cotdiem in mon.dhk1:
						for cotdiemhs in self.dhk1:
							if cotdiem.ten == cotdiemhs.ten:
								break
						else:
							tongheso+=cotdiem.heso
				else:
					for cotdiem in mon.dhk2:
						for cotdiemhs in self.dhk2:
							if cotdiem.ten == cotdiemhs.ten:
								break
						else:
							tongheso+=cotdiem.heso
		try:
			return round(tbmhk/tongheso,1)
		except ZeroDivisionError:
			return 0
class Diem:
	def __init__(self,Tendiem,*Heso):
		self.ten=Tendiem
		if Heso != ():
			self.heso=list(Heso)[0]
		else:
			self.heso=0
class Monhoc:
	def __init__(self,TenMonHoc):
		self.ten=TenMonHoc
		self.dhk1=[]
		self.dhk2=[]
		self.layout=None
		self.layout_hk1=None
		self.layout_hk2=None
		self.layout_cotdiem1=[]
		self.layout_cotdiem2=[]
		self.monchinh=False
class Chocsinh:
	def __init__(self, ten, NTNS):
		self.tmp=None
		self.diemhs=[]
		self.loi=[]
		self.ten=ten
		self.NTNS=NTNS
		self.truong=None
		self.phai=True#Nam
	def _phai(self):
		if self.phai:
			return ('human-male','Nam')
		else:
			return ('human-female','Nữ')
	def diemmonchinh(self,hocki):
		diemmonchinh=0
		monchinh=app.monhocchinh()
		for mon in self.diemhs:
			if mon.ten in monchinh:
				if diemmonchinh < mon.dtbmon(hocki):
					diemmonchinh = mon.dtbmon(hocki)
		return diemmonchinh
	def TBMthapnhat(self,hocki):
		tbmmin=None
		for mon in self.diemhs:
			try:
				if tbmmin < mon.dtbmon(hocki):
					tbmmin = mon.dtbmon(hocki)
			except TypeError:
				tbmmin = mon.dtbmon(hocki)
		return tbmmin
	def xeploai(self,hocki):
		TBcacmon = self.TBTCM(hocki)
		TBmonchinh = self.diemmonchinh(hocki)
		TBMmin = self.TBMthapnhat(hocki)
		if TBcacmon >= 8 and TBmonchinh >= 8 and TBMmin >= 6.5:
			return "Giỏi"
		elif TBcacmon >= 6.5 and TBmonchinh >= 6.5 and TBMmin >= 5:
			return "Khá"
		elif TBcacmon >= 5 and TBmonchinh >= 5 and TBMmin >= 3.5:
			return "Trung bình"
		elif TBcacmon >= 3.5 and TBMmin >= 2:
			return "Yếu"
		else:
			return "Kém"
	def themmh(self,TenMonHoc):
		self.diemhs.append(Diemhs(TenMonHoc))
	def TBTCM(self,hocki):
		TBTCM=0
		tongmonhoc=0
		if hocki==1:
			for mon in self.diemhs:
				if mon.dhk1 != []:
					TBTCM+=mon.dtbmon(1)
					tongmonhoc+=1
		else:
			for mon in self.diemhs:
				if mon.dhk2 != []:
					TBTCM+=mon.dtbmon(2)
					tongmonhoc+=1
		try:
			return round(TBTCM/tongmonhoc,1)
		except ZeroDivisionError:
			return 0
	def TBCaNamMH(self,monh):
		for mon in self.diemhs:
			if mon.ten == monh:
				if mon.dhk1 != [] and mon.dhk2 != []:
					return round((mon.dtbmon(1)+mon.dtbmon(2)*2)/3,1)
				else:
					return 0
	def TBCaNam(self):
		TBCN=0
		TongMonHoc=0
		for mon in self.diemhs:
			if mon.dhk1 != [] and mon.dhk2 != []:
				TBCN+=self.TBCaNamMH(mon.ten)
				TongMonHoc+=1
		return round(TBCN/TongMonHoc,1)
	def solanvipham(self,tenloi):
		sl=0
		for i in self.loi:
			if i.ten==tenloi:
				sl+=1
		return sl
	def tongtru(self,tenloi):
		return self.solanvipham(tenloi)*checkloi(tenloi).diem
	def tongdiemtru(self):
		diem=0
		for i in loi:
			diem+=self.tongtru(i.ten)
		return diem
class Clop:
	def __init__(self, ten):
		self.tshocsinh=0
		self.hocsinh=[]
		self.ten=ten
class Ctruong:
	def __init__(self, ten):
		self.lop=[]
		self.ten=ten
		self.tongslop=0





class layout_timeload(Screen):
	def __init__(self,**kwargn):
		super(layout_timeload,self).__init__(**kwargn)
		self.name="TimeLoad"
		self.add_widget(Image(
				source='loading.png',
				allow_stretch=True,
			))
class diemlayout(Screen):
	def __init__(self,*args,**kwargn):
		super(diemlayout,self).__init__(**kwargn)
		self.name="tabdiem"
		self.menubar=MDToolbar(
				title="Điểm học sinh", 
				type="top",
				pos_hint={'top':1},
				elevation=10,
		)
		self.tenhs=MDLabel(
				text=f"Tên học sinh:",
				color=(25/256,137/256,185/256,255/256),
				halign="center",
				pos_hint={'center_y':0.85},
				theme_text_color="ContrastParentBackground",
			)
		self.hocki_layout=Carousel(
				direction= 'right',
				size_hint_y=None,
				pos_hint={"center_x":.5,"center_y":.79},
				height=100
			)
		self.get_hocki=1
		self.hocki_layout.add_widget(MDLabel(
					text="Học kì I",
					font_style="H4",
					theme_text_color="Primary",
					halign="center"
				)
			)
		self.hocki_layout.add_widget(MDLabel(
					text="Học kì II",
					font_style="H4",
					theme_text_color="Primary",
					halign="center"
				)
			)
		with self.canvas.before:
			Color(rgba=(0, .4, 0, 0.1))
			self.rect=RoundedRectangle(radius=[(40.0, 40.0), (40.0, 40.0), (40.0, 40.0), (40.0, 40.0)])
		self.bind(pos=self.update_rect,size=self.update_rect)
		self.Sv=ScrollView(
			size_hint_y=.7,
			pos_hint={"x":0,"y":.06},
			do_scroll_x=False,
			do_scroll_y=True
		)
		self.Sv2=ScrollView(
			size_hint_y=.7,
			pos_hint={"x":0,"y":.06},
			do_scroll_x=False,
			do_scroll_y=True
		)
		self.View=GridLayout(
				size_hint_x=None,
				size_hint_y=None,
				cols=1,
				padding=(10, 10),
				spacing=10,
				size=(Window.width,Window.height)
			)
		self.View.bind(minimum_height=self.View.setter("height"))
		self.View2=GridLayout(
				size_hint_x=None,
				size_hint_y=None,
				cols=1,
				padding=(10, 10),
				spacing=10,
				size=(Window.width,Window.height)
			)
		self.View2.bind(minimum_height=self.View2.setter("height"))

		self.Sv.add_widget(self.View)
		self.Sv2.add_widget(self.View2)

		self.tab1=Screen(name="tab1")
		self.tab1.add_widget(self.Sv)
		self.tab2=Screen(name="tab2")
		self.tab2.add_widget(self.Sv2)
		self.tab=ScreenManager()
		self.tab.add_widget(self.tab1)
		self.tab.add_widget(self.tab2)
		self.hocki_layout.bind(index=self.hocki)

		self.add_widget(self.tab)
		self.add_widget(self.tenhs)
		self.add_widget(self.menubar)
		self.add_widget(self.hocki_layout)
		#BottomBar####################################################################################################################################################
		BottomBar=MDBottomAppBar()
		toolbar=MDToolbar(
				#title="Help",
				mode="end",
				icon="content-save",
			)
		toolbar.bind(on_action_button=lambda x:self.save(args[0]))
		#toolbar.left_action_items=[["coffee", lambda x: print("Main")]]
		toolbar.type="bottom"
		BottomBar.add_widget(toolbar)
		self.add_widget(BottomBar)
		##############################################################################################################################################################
	def update_rect(self,rect,a):
		rect.rect.pos=rect.pos
		rect.rect.size=(rect.size[0],rect.size[1]*83/100)
	def hocki(self,*args):
		if args[1]+1==1:
			self.tab.transition.direction="right"
			self.tab.current="tab1"
		else:
			self.tab.transition.direction="left"
			self.tab.current="tab2"
		self.get_hocki=args[1]+1
	def save(self,*args):
		app=args[0]
		for View in self.View.children[1:]:
			for View_children in View.children[0].children:
				for item in View_children.children:
					if hasattr(item, 'ten'):
						if item.ten == "diem":
							if item.children[1].error:
								thongbao(f"Lỗi: {item.children[1].helper_text}")
								return
		for View in self.View2.children[1:]:
			for View_children in View.children[0].children:
				for item in View_children.children:
					if hasattr(item, 'ten'):
						if item.ten == "diem":
							if item.children[1].error:
								thongbao(f"Lỗi: {item.children[1].helper_text}")
								return
		hs=app.geths(app.data_hocsinh[0],app.data_hocsinh[1],app.data_hocsinh[2],app.data_hocsinh[3])
		#Hocki1####################################################################################################################
		for View in self.View.children[1:]:
			if len(View.children[1].children[0].children)==1:
				tenmon=View.children[1].children[0].children[0].text[9:]
			else:
				tenmon=View.children[1].children[0].children[1].text[9:]
			diem=[]
			allmon=[]
			for mon in hs.diemhs:
				allmon.append(mon.ten)
			if tenmon not in allmon:
				hs.themmh(tenmon)
			for mon in hs.diemhs:
				if mon.ten == tenmon:
					for View_children in View.children[0].children:
						for item in View_children.children:
							if hasattr(item, 'ten'):
								if item.ten == "cotdiem":
									cotdiem=item.cot
									if mon.dhk1 == []:
										mon.dhk1.append(Diem(cotdiem))
									allcotdiem=[]
									for i in mon.dhk1:
										allcotdiem.append(i.ten)
									if cotdiem not in allcotdiem:
										mon.dhk1.append(Cotdiemhs(cotdiem))
									for cot in mon.dhk1:
										if cot.ten == cotdiem:
											cot.diem=diem
									diem=[]
								if item.ten == "diem":
									diem.append(item.children[1].text)
		#Hocki2#####################################################################################################################
		for View in self.View2.children[1:]:
			if len(View.children[1].children[0].children)==1:
				tenmon=View.children[1].children[0].children[0].text[9:]
			else:
				tenmon=View.children[1].children[0].children[1].text[9:]
			diem=[]
			allmon=[]
			for mon in hs.diemhs:
				allmon.append(mon.ten)
			if tenmon not in allmon:
				hs.themmh(tenmon)
			for mon in hs.diemhs:
				if mon.ten == tenmon:
					for View_children in View.children[0].children:
						for item in View_children.children:
							if hasattr(item, 'ten'):
								if item.ten == "cotdiem":
									cotdiem=item.cot
									if mon.dhk2 == []:
										mon.dhk2.append(Diem(cotdiem))
									allcotdiem=[]
									for i in mon.dhk2:
										allcotdiem.append(i.ten)
									if cotdiem not in allcotdiem:
										mon.dhk2.append(Cotdiemhs(cotdiem))
									for cot in mon.dhk2:
										if cot.ten == cotdiem:
											cot.diem=diem
									diem=[]
								if item.ten == "diem":
									diem.append(item.children[1].text)
		thongbao('Đã lưu điểm thành công')
		###############################################################################################################################
class layout_thietlapmon(GridLayout):
	def __init__(self,*args,**kwargn):
		super(layout_thietlapmon,self).__init__(**kwargn)
		self.Loading_Complate=False
		self.app=args[0]
		self.cols=1
		self.pos_hint={"center_x":.5,"center_y":.7}
		self.size_hint=(.8,.5)
		self.bind(pos=self.resize,size=self.resize)
		self.h1=GridLayout(cols=2)
		self.h2=GridLayout(cols=1,size_hint_y=None,height=140)
		self.add_widget(self.h1)
		self.add_widget(self.h2)
#-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*
		self.layout_hocki=Carousel(
				size_hint_x=None,
				width=120,
				loop=True,
				direction="bottom"
			)
		self.layout_hocki.bind(index=self.hocki)
		with self.layout_hocki.canvas.before:
			Color(rgba=(62/256, 112/256, 128/256, 1))
			self.layout_hocki.rect=RoundedRectangle(radius=[(40.0, 40.0), (0,0), (0,0), (0, 0)])
		self.layout_hocki.bind(pos=self.update_rect,size=self.update_rect)
		self.layout_hocki.add_widget(
				MDLabel(
					text="Học kì I",
					font_style="H6",
					theme_text_color="Primary",
					halign="center"
				)
			)
		self.layout_hocki.add_widget(
				MDLabel(
					text="Học kì II",
					font_style="H6",
					theme_text_color="Primary",
					halign="center"
				)
			)
		self.h1.add_widget(self.layout_hocki)

#-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*

		self.layout_thietlap=GridLayout(
				cols=1,
			)
		with self.layout_thietlap.canvas.before:
			Color(rgba=(101/256, 146/256, 159/256, 1))
			self.layout_thietlap.rect=RoundedRectangle(radius=[(0, 0), (40, 40), (0, 0), (0, 0)])
		self.layout_thietlap.bind(pos=self.update_rect,size=self.update_rect)
		
		#_*_*_*_*_*_*_*_*_*_*_*_*_*_*_*_*
		
		A=Screen(
				size_hint_y=None,
				height=100
			)
		self.Button_monchinh=MDFlatButton(
			text="Môn học chính",
			pos_hint={'center_x':.5,'center_y':.5},
		)
		self.Button_monchinh.bind(on_release=self.Monhocchinh)
		A.add_widget(self.Button_monchinh)
		self.layout_thietlap.add_widget(A)
		
		#_*_*_*_*_*_*_*_*_*_*_*_*_*_*_*_*
		
		B=ScrollView(
			do_scroll_x=False,
			do_scroll_y=True
		)
		self.thietlap_cotdiem=GridLayout(
				cols=3,
				size_hint_x=None,
				size_hint_y=None,
			)
		self.thietlap_cotdiem.bind(minimum_height=self.thietlap_cotdiem.setter("height"))
		B.add_widget(self.thietlap_cotdiem)
		self.layout_thietlap.add_widget(B)
		self.layout_thietlap.bind(size=self.resize_layout_thietlap)

		#_*_*_*_*_*_*_*_*_*_*_*_*_*_*_*_*

		C=Screen(size_hint_y=None,height=50)
		button=MDIconButton(icon="hospital",pos_hint={"center_x":.5,"center_y":.5})
		button.bind(on_press=self.add_cotdiem)
		C.add_widget(button)
		self.layout_thietlap.add_widget(C)
		self.h1.add_widget(self.layout_thietlap)
		
#-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*

		
		self.layout_monhoc=GridLayout(
				cols=2
			)
		self.list_monhoc=Carousel(
				direction="right",
				loop=True,
			)
		self.list_monhoc.bind(index=self.monhoc)
		tmp=Screen(size_hint_x=None,width=120)
		button=MDIconButton(icon="hospital",pos_hint={'center_x':.5,'center_y':.5})
		button.bind(on_press=self.add_monhoc)
		tmp.add_widget(button)
		self.layout_monhoc.add_widget(tmp)
		
		self.layout_monhoc.add_widget(self.list_monhoc)
		
		with self.layout_monhoc.canvas.before:
			Color(rgba=(62/256, 112/256, 128/256, 1))
			self.layout_monhoc.rect=RoundedRectangle(radius=[(0, 0), (0, 0), (40.0, 40.0), (40.0, 40.0)])
		self.layout_monhoc.bind(pos=self.update_rect,size=self.update_rect)


		self.h2.add_widget(self.layout_monhoc)

#-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*
		self.Loading_Complate=True
	def add_monhoc(self,*args):
		tmp=GridLayout(cols=2,size_hint_y=None,pos_hint={'top':1},)
		mon=MDTextField(
			text='',
			text_color=(153/256, 11/256, 164/256, 1),
			font_name="viet",
			font_size=sp(25),
			halign="center",
			hint_text="Môn",
			helper_text="",
			helper_text_mode="persistent",
		)
		self.app.thietlapmon_xong.mon_TextField.append(mon)
		mon.bind(text=self.nhap_tencot_tenmon)
		mon.bind(focus=self.on_slec)
		mon.ten="tenmon"
		button=MDIconButton(icon="close",size_hint_x=None,width=140)
		button.mon=mon
		button.bind(on_press=self.xoamonhoc)
		tmp.add_widget(mon)
		tmp.add_widget(button)
		self.list_monhoc.add_widget(tmp)
		self.app.monhoc.append(Monhoc(''))
		self.list_monhoc.load_slide(self.list_monhoc.slides[-1])
		self.mon_slec=self.app.monhoc[-1]
		self.mon_slec.monchinh=False
		self.Button_monchinh.text="Môn học phụ"
		self.mon_slec_old=self.mon_slec
		#Thiet lap layout mới#################################################################
		View=GridLayout(
				rows=1,
				size_hint_y=None,
			)
		col1=GridLayout(
			size_hint_x=None,
			width=210,
			cols=1,
		)
		View.add_widget(col1)
		col2=GridLayout(
			cols=1,
		)
		self.mon_slec.col2hk1=col2
		col2.bind(size=self.app.resize_col2)
		View.add_widget(col2)
		View1=GridLayout(
			cols=1,
		)
		monhoc=MDLabel(text=f"Môn Học:\n{self.mon_slec.ten}",halign="center",theme_text_color="Custom",text_color=(102, 255, 102))
		self.mon_slec.layout_mon1=monhoc
		View1.add_widget(monhoc)
		col1.add_widget(View1)
		col2.bind(children=self.app.resize_diem)
		self.app.resize_diem(col2,col2.children)
		for cot in self.mon_slec.dhk2:
			pass
		with View.canvas.before:
			Color(rgba=(0, .4, 0, 0.2))
			View.rect=RoundedRectangle(radius=[(40.0, 40.0), (40.0, 40.0), (40.0, 40.0), (40.0, 40.0)])
		View.bind(pos=self.app.update_rect,size=self.app.update_rect)
		self.mon_slec.layout_hk1=View
		button.layout_hk1=View
		#Hocki2##########################################################################################################################################################
		View=GridLayout(
				rows=1,
				size_hint_y=None,
			)
		col1=GridLayout(
			size_hint_x=None,
			width=210,
			cols=1,
		)
		View.add_widget(col1)
		col2=GridLayout(
			cols=1,
		)
		self.mon_slec.col2hk2=col2
		col2.bind(size=self.app.resize_col2)
		View.add_widget(col2)
		View1=GridLayout(
			cols=1,
		)
		monhoc=MDLabel(text=f"Môn Học:\n{self.mon_slec.ten}",halign="center",theme_text_color="Custom",text_color=(102, 255, 102))
		self.mon_slec.layout_mon2=monhoc
		View1.add_widget(monhoc)
		col1.add_widget(View1)
		col2.bind(children=self.app.resize_diem)
		self.app.resize_diem(col2,col2.children)
		with View.canvas.before:
			Color(rgba=(0, .4, 0, 0.2))
			View.rect=RoundedRectangle(radius=[(40.0, 40.0), (40.0, 40.0), (40.0, 40.0), (40.0, 40.0)])
		View.bind(pos=self.app.update_rect,size=self.app.update_rect)
		self.mon_slec.layout_hk2=View
		button.layout_hk2=View
	def add_cotdiem(self,*args):
		if self.list_monhoc.index==None:return
		cd=MDTextField(
			text="",
			pos_hint={'top':.6},
			font_name="viet",
			font_size=sp(15),
			text_color=(153/256, 11/256, 164/256, 1),
			halign="center",
			hint_text="Cột điểm",
			helper_text="",
			helper_text_mode="persistent",
		)
		self.app.thietlapmon_xong.cotdiem_TextField.append(cd)
		cd.ten="cotdiem"
		cd.bind(focus=self.on_slec)
		cd.bind(text=self.nhap_tencot_tenmon)
		if self.hocki_var == 1:
			self.mon_slec.layout_cotdiem1.append(cd)
		else:
			self.mon_slec.layout_cotdiem2.append(cd)
		self.thietlap_cotdiem.add_widget(cd)
		heso=MDTextField(
			text="",
			font_name="viet",
			pos_hint={'top':.6},
			font_size=sp(15),
			text_color=(181/256, 157/256, 5/256, 1),
			halign="center",
			hint_text="Hệ số",
			helper_text="",
			helper_text_mode="persistent",
		)
		self.app.thietlapmon_xong.heso_TextField.append(heso)
		heso.ten="heso"
		heso.bind(focus=self.on_slec_heso)
		heso.bind(text=self.nhap_heso)
		if self.hocki_var == 1:
			self.mon_slec.layout_cotdiem1.append(heso)
		else:
			self.mon_slec.layout_cotdiem2.append(heso)
		self.thietlap_cotdiem.add_widget(heso)

		button=MDIconButton(icon="close",size_hint_x=None,width=10)
		button.cotdiem=cd
		button.heso=heso
		button.bind(on_press=self.xoacotdiem)
		self.thietlap_cotdiem.add_widget(button)

		if self.hocki_var==1:
			self.mon_slec.dhk1.append(Diem(""))
			cd.cot=self.mon_slec.dhk1[-1]
			button.layout1=cd.cot
			heso.cotdiem_pos=cd.cot
		else:
			self.mon_slec.dhk2.append(Diem(""))
			cd.cot=self.mon_slec.dhk2[-1]
			button.layout2=cd.cot
			heso.cotdiem_pos=cd.cot

		cotdiem=MDLabel(text=f"Cột Điểm: {cd.cot.ten}",halign="center",theme_text_color="Custom",text_color=(51/256, 153/256, 102/256))
		cd.cot.layout_cot=cotdiem
		cotdiem.ten="cotdiem"
		cotdiem.cot=cd.cot.ten
		View2=GridLayout(
			cols=1,
			row_default_height= 50,
			row_force_default=True,
			size_hint_y=None,
			height=40
		)
		View2.add_widget(cotdiem)
		all_diem=GridLayout(cols=1,size_hint_y=None,)
		all_diem.cot=cd.cot
		all_diem.ten="alldiem"
		all_diem.bind(children=self.app.resize_alldiem)
		cd.cot.delete=[all_diem,View2]
		layout=Screen(size_hint_y=None,size_hint_x=None,height=50*2)
		add_button=MDIconButton(icon="hospital",pos_hint={'center_x':.5,'center_y':.5})
		add_button.bind(on_press=self.app.themdiem)
		layout.add_widget(add_button)
		layout.ten="adddiem"
		if self.hocki_var==1:
			self.mon_slec.col2hk1.add_widget(View2)
			self.mon_slec.col2hk1.add_widget(all_diem)
		else:
			self.mon_slec.col2hk2.add_widget(View2)
			self.mon_slec.col2hk2.add_widget(all_diem)
		all_diem.add_widget(layout)
		layout.width=View2.parent.width-100

		self.nhap_heso(heso)
		self.nhap_tencot_tenmon(cd)
	def hocki(self,*args):
		if self.Loading_Complate:
			self.hocki_var=args[1]+1
			if self.list_monhoc.children != []:self.monhoc_var=self.list_monhoc.children[0].children[0].children[1].text
			if hasattr(self,"monhoc_var"):
				self.load_cotdiem(self.hocki_var,self.monhoc_var)
	def monhoc(self,*args):
		if self.Loading_Complate:
			self.monhoc_var=args[0].children[0].children[0].children[1].text
			self.hocki_var=self.layout_hocki.index+1
			if hasattr(self,"hocki_var"):
				self.load_cotdiem(self.hocki_var,self.monhoc_var)
	def Monhocchinh(self,*args):
		if self.list_monhoc.index==None:return
		if args[0].text == "Môn học chính":
			args[0].text = "Môn học phụ"
			args[0].mon.monchinh=False
		else:
			args[0].text = "Môn học chính"
			args[0].mon.monchinh=True
	def resize(self,*args):
		window=args[0]
		self.app.thietlapmon_xong.pos=(window.pos[0]+window.width-self.app.thietlapmon_xong.width,window.pos[1]+self.height-self.app.thietlapmon_xong.size[1])
	def resize_layout_thietlap(self,*args):
		self.thietlap_cotdiem.width=args[1][0]
	def update_rect(self,rect,a):
		rect.rect.pos=rect.pos
		rect.rect.size=rect.size
	def update_rect_(self,rect,a):
		rect.rect.pos=(rect.pos[0]-50,rect.pos[1]-50)
		rect.rect.size=(rect.size[0]+100,rect.size[1]+100)
	def open(self,*args):
		try:
			self.list_monhoc.clear_widgets()
		except:
			pass
		for i in self.app.monhoc:
			tmp=GridLayout(cols=2,size_hint_y=None,pos_hint={'top':1},)
			mon=MDTextField(
				font_name="viet",
				text=i.ten,
				text_color=(153/256, 11/256, 164/256, 1),
				font_size=sp(25),
				halign="center",
				hint_text="Môn",
				helper_text="",
				helper_text_mode="persistent",
				)
			self.app.thietlapmon_xong.mon_TextField.append(mon)
			mon.bind(text=self.nhap_tencot_tenmon)
			mon.bind(focus=self.on_slec)
			mon.ten="tenmon"
			button=MDIconButton(icon="close",size_hint_x=None,width=140)
			button.mon=mon
			button.bind(on_press=self.xoamonhoc)
			button.layout_hk1=i.layout_hk1
			button.layout_hk2=i.layout_hk2
			tmp.add_widget(mon)
			tmp.add_widget(button)
			self.list_monhoc.add_widget(tmp)
	def xoamonhoc(self,*args):
		self.app.monhoc.remove(self.mon_slec)
		button=args[0]
		self.app.thietlapmon_xong.mon_TextField.remove(button.mon)
		for i in self.thietlap_cotdiem.children:
			if type(i)==kivymd.uix.textfield.MDTextField:
				if i.hint_text=="Cột điểm":
					self.app.thietlapmon_xong.cotdiem_TextField.remove(i)
				else:
					self.app.thietlapmon_xong.heso_TextField.remove(i)
		button.parent.parent.parent.remove_widget(button.parent)
		if self.list_monhoc.children==[]:
			self.monhoc_var=""
			self.Button_monchinh.text=""
			self.load_cotdiem(self.hocki_var,self.monhoc_var)
		else:
			self.monhoc_var=self.list_monhoc.children[0].children[0].children[1].text
			self.load_cotdiem(self.hocki_var,self.monhoc_var)
		if button.layout_hk1.parent != None:button.layout_hk1.parent.remove_widget(button.layout_hk1)
		if button.layout_hk2.parent != None:button.layout_hk2.parent.remove_widget(button.layout_hk2)
	def on_slec(self,*args):
		if args[0].text=="" and args[1]==True:
			if args[0].ten=="cotdiem":
				args[0].helper_text="Tên cột điểm không thể bỏ trống"
			else:
				args[0].helper_text="Tên môn không thể bỏ trống"
		else:
			args[0].helper_text=""
	def on_slec_heso(self,*args):
		try:
			diem=float(args[0].text)
			if diem <= 0:
				args[0].helper_text="Hệ số phải lớn hơn 0"
				args[0].error=True
				args[0]._anim_current_line_color(args[0].error_color)
			else:
				args[0].helper_text=""
		except:
			if args[1]==True:
				if args[0].text=="":
					args[0].helper_text="Hệ số không thể bỏ trống"
				else:
					args[0].helper_text="Kí tự này không phải là hệ số"
			else:
				args[0].helper_text=""
	def load_cotdiem(self,hocki,mon):
		self.thietlap_cotdiem.clear_widgets()
		for i in self.app.monhoc:
			if i.ten == mon:
				if i.monchinh:
					self.Button_monchinh.text="Môn học chính"
				else:
					self.Button_monchinh.text="Môn học phụ"
				self.Button_monchinh.mon=i
				tmp=False
				if hocki==1:
					for layout in i.layout_cotdiem1:
						self.thietlap_cotdiem.add_widget(layout)
						if tmp:
							button.heso=layout
							self.thietlap_cotdiem.add_widget(button)
							tmp=False
						else:
							button=MDIconButton(icon="close",size_hint_x=None,width=10)
							button.bind(on_press=self.xoacotdiem)
							button.cotdiem=layout
							tmp=True
				if hocki==2:
					for layout in i.layout_cotdiem2:
						self.thietlap_cotdiem.add_widget(layout)
						if tmp:
							button.heso=layout
							self.thietlap_cotdiem.add_widget(button)
							tmp=False
						else:
							button=MDIconButton(icon="close",size_hint_x=None,width=10)
							button.bind(on_press=self.xoacotdiem)
							button.cotdiem=layout
							tmp=True
				if hasattr(self,"mon_slec"):
					self.mon_slec_old=self.mon_slec
					self.mon_slec=i
				else:
					self.mon_slec=i
					self.mon_slec_old=i
	def xoacotdiem(self,*args):
		parent=args[0].parent
		self.app.thietlapmon_xong.cotdiem_TextField.remove(args[0].cotdiem)
		self.app.thietlapmon_xong.heso_TextField.remove(args[0].heso)
		if self.hocki_var == 1:
			self.mon_slec.layout_cotdiem1.remove(args[0].cotdiem)
			self.mon_slec.layout_cotdiem1.remove(args[0].heso)
			self.mon_slec.dhk1.remove(args[0].cotdiem.cot)
			args[0].cotdiem.cot.layout_cot.parent.parent.remove_widget(args[0].cotdiem.cot.delete[0])
			args[0].cotdiem.cot.layout_cot.parent.parent.remove_widget(args[0].cotdiem.cot.delete[1])
		else:
			self.mon_slec.layout_cotdiem2.remove(args[0].cotdiem)
			self.mon_slec.layout_cotdiem2.remove(args[0].heso)
			self.mon_slec.dhk2.remove(args[0].cotdiem.cot)
			args[0].cotdiem.cot.layout_cot.parent.parent.remove_widget(args[0].cotdiem.cot.delete[0])
			args[0].cotdiem.cot.layout_cot.parent.parent.remove_widget(args[0].cotdiem.cot.delete[1])
		parent.remove_widget(args[0].cotdiem)
		parent.remove_widget(args[0].heso)
		parent.remove_widget(args[0])
	def load_all_cotdiem(self):
		for i in self.app.monhoc:
			i.layout_cotdiem1=[]
			i.layout_cotdiem2=[]
			for cotdiem in i.dhk1:
				cd=MDTextField(
					font_name="viet",
					text=str(cotdiem.ten),
					pos_hint={'top':.6},
					font_size=sp(15),
					text_color=(153/256, 11/256, 164/256, 1),
					halign="center",
					helper_text="",
					hint_text="Cột điểm",
					helper_text_mode="persistent",
				)
				self.app.thietlapmon_xong.cotdiem_TextField.append(cd)
				cd.cot=cotdiem
				cd.ten="cotdiem"
				cd.bind(text=self.nhap_tencot_tenmon)
				cd.bind(focus=self.on_slec)
				i.layout_cotdiem1.append(cd)
				heso=MDTextField(
					font_name="viet",
					text=str(cotdiem.heso),
					pos_hint={'top':.6},
					font_size=sp(15),
					text_color=(181/256, 157/256, 5/256, 1),
					halign="center",
					hint_text="Hệ số",
					helper_text="",
					helper_text_mode="persistent",
				)
				self.app.thietlapmon_xong.heso_TextField.append(heso)
				heso.cotdiem_pos=cotdiem
				heso.ten="heso"
				heso.bind(text=self.nhap_heso)
				heso.bind(focus=self.on_slec_heso)
				i.layout_cotdiem1.append(heso)
			for cotdiem in i.dhk2:
				cd=MDTextField(
					font_name="viet",
					text=str(cotdiem.ten),
					pos_hint={'top':.6},
					font_size=sp(15),
					halign="center",
					hint_text="Cột điểm",
					text_color=(153/256, 11/256, 164/256, 1),
					helper_text="",
					helper_text_mode="persistent",
				)
				self.app.thietlapmon_xong.cotdiem_TextField.append(cd)
				cd.cot=cotdiem
				cd.ten="cotdiem"
				cd.bind(focus=self.on_slec)
				cd.bind(text=self.nhap_tencot_tenmon)
				i.layout_cotdiem2.append(cd)
				heso=MDTextField(
					font_name="viet",
					text=str(cotdiem.heso),
					pos_hint={'top':.6},
					font_size=sp(15),
					halign="center",
					hint_text="Hệ số",
					text_color=(181/256, 157/256, 5/256, 1),
					helper_text="",
					helper_text_mode="persistent",
				)
				self.app.thietlapmon_xong.heso_TextField.append(heso)
				heso.cotdiem_pos=cotdiem
				heso.ten="heso"
				heso.bind(focus=self.on_slec_heso)
				heso.bind(text=self.nhap_heso)
				i.layout_cotdiem2.append(heso)
	def nhap_heso(self,*args):
		try:
			diem=float(args[0].text)
			if diem <= 0:
				args[0].helper_text="Hệ số phải lớn hơn 0"
				args[0].error_text="Hệ số phải lớn hơn 0"
				args[0].error=True
				args[0].cotdiem_pos.heso=0
				args[0]._anim_current_line_color(args[0].error_color)
			else:
				args[0].cotdiem_pos.heso=diem
				args[0].helper_text=""
				args[0].error=False
				args[0]._anim_current_line_color(args[0].theme_cls.disabled_hint_text_color)
		except:
			args[0].cotdiem_pos.heso=0
			if args[0].text=="":
				args[0].helper_text="Hệ số không thể bỏ trống"
				args[0].error_text="Hệ số không thể bỏ trống"
			else:
				args[0].helper_text="Kí tự này không phải là hệ số"
				args[0].error_text="Kí tự này không phải là hệ số"
			args[0].error=True
			args[0]._anim_current_line_color(args[0].error_color)
	def nhap_tencot_tenmon(self,*args):
		if args[0].text=="":
			if args[0].ten=="cotdiem":
				args[0].helper_text="Tên cột điểm không thể bỏ trống"
				args[0].error_text="Tên cột điểm không thể bỏ trống"
			else:
				args[0].error_text="Tên môn không thể bỏ trống"
				args[0].helper_text="Tên môn không thể bỏ trống"
			args[0].error=True
			args[0]._anim_current_line_color(args[0].error_color)
		else:
			args[0].helper_text=""
			args[0].error=False
			args[0]._anim_current_line_color(args[0].theme_cls.disabled_hint_text_color)
		if args[0].ten=='tenmon':
			self.mon_slec.ten=args[0].text
			self.mon_slec.layout_mon1.text=f"Môn Học:\n{args[0].text}"
			self.mon_slec.layout_mon2.text=f"Môn Học:\n{args[0].text}"
		if args[0].ten=='cotdiem':
			args[0].cot.ten=args[0].text
			args[0].cot.layout_cot.text=f"Cột Điểm: {args[0].text}"
			args[0].cot.layout_cot.cot=args[0].text
class Layout_Chart(Screen):
	def __init__(self,*args,**kwargn):
		super(Layout_Chart,self).__init__(**kwargn)
		self.name="Chart"
		font = {'family' : 'normal',
        'weight' : 'bold',
        'size'   : 22}
		matplotlib.rc('font', **font)
		
		self.menubar=MDToolbar(
				title="Thống kê dữ liệu", 
				type="top",
				pos_hint={'top':1},
				elevation=10,
		)
		self.menubar.left_action_items=[["backspace", lambda x: app.move_tab("Main")]]
		self.add_widget(self.menubar)

		self.MDTabs=MDTabs(size_hint=(1,None))
		self.MDTabs.height=Window.height-self.menubar.height
		self.MDTabs.bind(on_tab_switch=self.slec_tab)
		self.add_widget(self.MDTabs)

		self.tabs1=MDTabsBase(title="Kết quả học tập")
		self.ketqua=GridLayout(cols=1,padding=20,size=(self.MDTabs.size[0],self.MDTabs.size[1]-100))
		self.tabs1.add_widget(self.ketqua)
		self.button_ketqua=GridLayout(rows=1,size_hint=(None,None))
		self.hocki_ketqua=MDRectangleFlatButton(text="Học kì I")
		self.hocki_ketqua.bind(on_release=self.change_hocki)
		self.button_ketqua.add_widget(self.hocki_ketqua)
		self.lop_ketqua=MDRectangleFlatButton(text="Toàn khối")
		self.lop_ketqua.bind(on_release=self.open_lop_ketqua)
		self.button_ketqua.add_widget(self.lop_ketqua)
		self.ketqua.add_widget(self.button_ketqua)
		self.MDTabs.add_widget(self.tabs1)


		self.diemcacmon=MDTabsBase(title="Điểm")
		self.MDTabs.add_widget(self.diemcacmon)
		self.diemcacmon_layout=GridLayout(cols=1,padding=20,size=(self.MDTabs.size[0],self.MDTabs.size[1]-100))
		self.all_button=GridLayout(rows=1,size_hint=(None,None))
		self.hocki=MDRectangleFlatButton(text="Học kì I")
		self.hocki.bind(on_release=self.change_hocki_diem)
		self.monhoc=MDRectangleFlatButton(text="Tất cả môn")
		self.monhoc.bind(on_release=self.open_menu_monhoc)
		self.cotdiem=MDRectangleFlatButton(text="Tất cả cột")
		self.cotdiem.bind(on_release=self.open_menu_cotdiem)
		self.lop=MDRectangleFlatButton(text="Toàn khối")
		self.lop.bind(on_release=self.open_menu_lop)
		self.all_button.add_widget(self.hocki)
		self.all_button.add_widget(self.lop)
		self.all_button.add_widget(self.monhoc)
		self.all_button.add_widget(self.cotdiem)
		self.diemcacmon_layout.add_widget(self.all_button)
		self.diemcacmon.add_widget(self.diemcacmon_layout)

		self.MDTabs.bind(size=self.resize)
	def slec_tab(self,*args):
		try:
			if args[3]=="Kết quả học tập" and "Kết quả học tập" in self.all_tab:
				self.update_chart()
				self.all_tab.remove("Kết quả học tập")
			elif args[3]=="Điểm" and 'Điểm' in self.all_tab:
				self.update_diem()
				self.all_tab.remove('Điểm')
		except:
			self.lop_ketqua.text="Toàn khối"
			self.monhoc.text="Tất cả môn"
			self.cotdiem.text="Tất cả cột"
			self.lop.text="Toàn khối"
			self.all_tab=['Điểm','Kết quả học tập']
			if self.MDTabs.carousel.current_slide.tab_label.text == "Kết quả học tập":
				self.update_chart()
				self.all_tab.remove("Kết quả học tập")
			elif self.MDTabs.carousel.current_slide.tab_label.text == "Điểm":
				self.update_diem()
				self.all_tab.remove('Điểm')
	def update_diem(self):
		pl.clf()
		pl.tight_layout(pad=5)

		self.diemcacmon_chart()
	def diemcacmon_chart(self):
		all_diem=[]
		for tr in app.truong:
			for lop in tr.lop:
				if lop.ten == self.lop.text or self.lop.text == "Toàn khối":
					for hs in lop.hocsinh:
						for mon in hs.diemhs:
							if mon.ten == self.monhoc.text or self.monhoc.text == "Tất cả môn":
								if self.hocki.text=="Học kì I":
									for cotdiem in mon.dhk1:
										if cotdiem.ten == self.cotdiem.text or self.cotdiem.text == "Tất cả cột":
											for diem in cotdiem.diem:
												all_diem.append(diem)
								else:
									for cotdiem in mon.dhk2:
										if cotdiem.ten == self.cotdiem.text or self.cotdiem.text == "Tất cả cột":
											for diem in cotdiem.diem:
												all_diem.append(diem)
		data1=[]
		data2=[]
		all_diem.sort()
		annotate=[]
		countt=0
		for i in all_diem:
			if i not in data2:
				data1.append(all_diem.count(i))
				annotate.append([all_diem.count(i),(str(i),all_diem.count(i))])
				data2.append(str(i))
				countt+=1
		if len(self.diemcacmon_layout.children)==2:
			self.diemcacmon_layout.remove_widget(self.diemcacmon_layout.children[0])
		pl.clf()
		pl.tight_layout(pad=5)
		pl.bar(data2,data1,color="#81F0D1")
		pl.xlabel("Điểm")
		pl.ylabel("Số học sinh")
		for i in annotate:
			pl.annotate(
				i[0],
				xy=i[1]
				)
		self.chart1=FigureCanvasKivyAgg(pl.gcf())
		self.diemcacmon_layout.add_widget(self.chart1)
	def change_hocki_diem(self,*args):
		if args[0].text == "Học kì I":
			args[0].text="Học kì II"
		else:
			args[0].text="Học kì I"
		self.cotdiem.text="Tất cả cột"
		self.diemcacmon_chart()
	def open_menu_lop(self,*args):
		self.menu=MDDropdownMenu(
				caller=args[0],
				max_height=500,
				width_mult=4,
				position="bottom"
			)
		all_lop=[]
		if self.lop.text != "Toàn khối":
			self.menu.items.append({
							"text":"Toàn khối",
							"viewclass": "OneLineListItem",
							"on_release":lambda x="Toàn khối":self.slec_lop(x),
						})
		for tr in app.truong:
			for lop in tr.lop:
				if lop.ten not in all_lop and lop.ten != self.lop.text:
					self.menu.items.append({
							"text":lop.ten,
							"viewclass": "OneLineListItem",
							"on_release":lambda x=lop.ten:self.slec_lop(x),
						})
					all_lop.append(lop.ten)
		if self.menu.items!=[]:self.menu.open()
	def open_lop_ketqua(self,*args):
		self.menu=MDDropdownMenu(
				caller=args[0],
				max_height=500,
				width_mult=4,
				position="bottom"
			)
		all_lop=[]
		if self.lop_ketqua.text != "Toàn khối":
			self.menu.items.append({
							"text":"Toàn khối",
							"viewclass": "OneLineListItem",
							"on_release":lambda x="Toàn khối":self.slec_lop_ketqua(x),
						})
		for tr in app.truong:
			for lop in tr.lop:
				if lop.ten not in all_lop and lop.ten != self.lop_ketqua.text:
					self.menu.items.append({
							"text":lop.ten,
							"viewclass": "OneLineListItem",
							"on_release":lambda x=lop.ten:self.slec_lop_ketqua(x),
						})
					all_lop.append(lop.ten)
		if self.menu.items!=[]:self.menu.open()
	def slec_lop_ketqua(self,lop):
		self.lop_ketqua.text=lop
		self.update_chart()
		self.menu.dismiss()
	def open_menu_monhoc(self,*args):
		self.menu=MDDropdownMenu(
				caller=args[0],
				max_height=500,
				width_mult=4,
				position="bottom"
			)
		if self.monhoc.text != "Tất cả môn":
			self.menu.items.append({
							"text":"Tất cả môn",
							"viewclass": "OneLineListItem",
							"on_release":lambda x="Tất cả môn":self.slec_mon(x),
						})
		for mon in app.monhoc:
			if mon.ten != self.monhoc.text:
				self.menu.items.append({
						"text":mon.ten,
						"viewclass": "OneLineListItem",
						"on_release":lambda x=mon.ten:self.slec_mon(x),
					})
		if self.menu.items!=[]:self.menu.open()
	def slec_lop(self,lop):
		self.lop.text=lop
		self.menu.dismiss()
		self.diemcacmon_chart()
	def slec_mon(self,mon):
		self.monhoc.text=mon
		self.cotdiem.text="Tất cả cột"
		self.menu.dismiss()
		self.diemcacmon_chart()
	def slec_cotdiem(self,cotdiem):
		self.cotdiem.text=cotdiem
		self.menu.dismiss()
		self.diemcacmon_chart()
	def open_menu_cotdiem(self,*args):
		self.menu=MDDropdownMenu(
				caller=args[0],
				max_height=500,
				width_mult=4,
				position="bottom"
			)
		if self.cotdiem.text != "Tất cả cột":
			self.menu.items.append({
							"text":"Tất cả cột",
							"viewclass": "OneLineListItem",
							"on_release":lambda x="Tất cả cột":self.slec_cotdiem(x),
						})
		all_cot=[]
		for mon in app.monhoc:
			if mon.ten == self.monhoc.text or self.monhoc.text == "Tất cả môn":
				if self.hocki.text=="Học kì I":
					for cotdiem in mon.dhk1:
						if cotdiem.ten not in all_cot:
							self.menu.items.append({
									"text":cotdiem.ten,
									"viewclass": "OneLineListItem",
									"on_release":lambda x=cotdiem.ten:self.slec_cotdiem(x),
								})
							all_cot.append(cotdiem.ten)
				else:
					for cotdiem in mon.dhk2:
						if cotdiem.ten not in all_cot:
							self.menu.items.append({
									"text":cotdiem.ten,
									"viewclass": "OneLineListItem",
									"on_release":lambda x=cotdiem.ten:self.slec_cotdiem(x),
								})
							all_cot.append(cotdiem.ten)
		if self.menu.items!=[]:self.menu.open()
	def change_hocki(self,*args):
		self.fig.clf()
		self.fig, self.axs = pl.subplots(2)
		self.fig.tight_layout(pad=5)
		if args[0].text=="Học kì I":
			self.bar_chart(2)
			self.pie_chart(2)
			args[0].text="Học kì II"
		else:
			self.bar_chart(1)
			self.pie_chart(1)
			args[0].text="Học kì I"
		self.ketqua.remove_widget(self.chart)
		self.chart=FigureCanvasKivyAgg(pl.gcf())
		self.ketqua.add_widget(self.chart)
	def update_chart(self):
		if len(self.ketqua.children)==2:
			self.ketqua.remove_widget(self.ketqua.children[0])
		if hasattr(self,"fig"):
			self.fig.clf()

		self.fig, self.axs = pl.subplots(2)
		self.fig.tight_layout(pad=5)
		if self.hocki_ketqua.text=="Học kì I":
			self.bar_chart(1)
			self.pie_chart(1)
		else:
			self.bar_chart(2)
			self.pie_chart(2)

		self.chart=FigureCanvasKivyAgg(pl.gcf())
		self.ketqua.add_widget(self.chart)
	def bar_chart(self,hocki):
		xeploai=[]
		label=[]
		annotate=[]
		tonghs_gioi=app.tong_hocsinh_gioi(hocki,self.lop_ketqua.text)
		tonghs_kha=app.tong_hocsinh_kha(hocki,self.lop_ketqua.text)
		tonghs_trungbinh=app.tong_hocsinh_trungbinh(hocki,self.lop_ketqua.text)
		tonghs_yeu=app.tong_hocsinh_yeu(hocki,self.lop_ketqua.text)
		tonghs_kem=app.tong_hocsinh_kem(hocki,self.lop_ketqua.text)
		color=[]
		if tonghs_gioi!=0:
			xeploai.append(tonghs_gioi)
			label.append("Giỏi")
			color.append("#81F0D1")
			annotate.append([tonghs_gioi,(len(label)-1,tonghs_gioi)])
		if tonghs_kha!=0:
			xeploai.append(tonghs_kha)
			label.append("Khá")
			color.append("#ACF081")
			annotate.append([tonghs_kha,(len(label)-1,tonghs_kha)])
		if tonghs_trungbinh!=0:
			xeploai.append(tonghs_trungbinh)
			label.append("Trung bình")
			color.append("#D8F081")
			annotate.append([tonghs_trungbinh,(len(label)-1,tonghs_trungbinh)])
		if tonghs_yeu!=0:
			xeploai.append(tonghs_yeu)
			label.append("Yếu")
			color.append("#F0B581")
			annotate.append([tonghs_yeu,(len(label)-1,tonghs_yeu)])
		if tonghs_kem!=0:
			xeploai.append(tonghs_kem)
			label.append("Kém")
			color.append("#FF7272")
			annotate.append([tonghs_kem,(len(label)-1,tonghs_kem)])
		self.axs[1].bar(label,xeploai,color=color)
		for i in annotate:
			self.axs[1].annotate(
				i[0],
				xy=i[1]
				)
	def pie_chart(self,hocki):
		#Chuẩn bị dữ liệu
		gioi=app.tong_hocsinh_gioi(hocki,self.lop_ketqua.text)
		kha=app.tong_hocsinh_kha(hocki,self.lop_ketqua.text)
		trungbinh=app.tong_hocsinh_trungbinh(hocki,self.lop_ketqua.text)
		yeu=app.tong_hocsinh_yeu(hocki,self.lop_ketqua.text)
		kem=app.tong_hocsinh_kem(hocki,self.lop_ketqua.text)
		tongsohs=gioi+kha+trungbinh+yeu+kem
		data=[]
		data1=[]
		explode=[]
		color=[]
		if gioi != 0:
			data.append(gioi/tongsohs*100)
			data1.append("Giỏi")
			explode.append(0)
			color.append("#81F0D1")
		if kha != 0:
			data.append(kha/tongsohs*100)
			data1.append("Khá")
			explode.append(0)
			color.append("#ACF081")
		if trungbinh != 0:
			data.append(trungbinh/tongsohs*100)
			data1.append("Trung bình")
			explode.append(0)
			color.append("#D8F081")
		if yeu != 0:
			data.append(yeu/tongsohs*100)
			data1.append("Yếu")
			explode.append(0)
			color.append("#F0B581")
		if kem != 0:
			data.append(kem/tongsohs*100)
			data1.append("Kém")
			explode.append(0)
			color.append("#FF7272")
		if gioi == 0 and kha == 0 and trungbinh == 0 and yeu == 0 and kem == 0:
			return
		index=0
		all_index=[]
		for i in data:
			if i==min(data):
				all_index.append(index)
			index+=1
		for index in all_index:
			explode[index]=0.1
		self.axs[0].pie(
				data,
				labels=data1,
				explode=explode,
				autopct="%1.3f%%",
				colors=color,
				wedgeprops={"edgecolor":"white","linewidth":2},
			)
	def resize(self,*args):
		self.diemcacmon_layout.size=(self.MDTabs.size[0],self.MDTabs.size[1]-100)
		self.ketqua.size=(self.MDTabs.size[0],self.MDTabs.size[1]-100)
class Layout_Hocsinh(Screen):
	def __init__(self,*args,**kwargn):
		super(Layout_Hocsinh,self).__init__(**kwargn)
		self.app=args[0]
		self.datatb=DT(
			rows_num=6,
			use_pagination=True,
			pagination_menu_height = "240dp",
			pagination_menu_pos = 'auto',
			# center_x=0,
			size_hint=(1,.7),
			check=True,
			pos_hint={'center_y':.53},
			elevation=0,
			column_data=[
				("Mã-HS",dp(35)),
				("Lớp",dp(30)),
				("Tên học sinh",dp(30)),
				("NTNS",dp(25)),
				("Phái",dp(15))
			],
		)
		self.datatb.pagination.children[-1].text="Tổng học sinh trong một trang"
		self.datatb.indexx=[]
		self.datatb.bind(on_check_press=self.app.on_check_press)
		self.datatb.header.ids.check.parent.remove_widget(self.datatb.header.ids.check)
		for i in self.datatb.header.children:
			for j in i.children:
				if type(j) == kivymd.uix.datatables.CellHeader:
					j.tooltip_text=""
		self.datatb.header.ids.first_cell.tooltip_text=""
		self.datatb.bind(on_row_press=self.app.selec_row)
		self.add_widget(self.datatb)
		self.Action_Button=MDFloatingActionButtonSpeedDial(
				root_button_anim= True,
				data={
					"Thống kê":"chart-bar",
					"Thêm học sinh":"account-plus",
					"Xóa học sinh":"account-minus",
					"Xuất dữ liệu điểm":"database",
					"Lưu dữ liệu":"content-save",
					"Lưu dữ liệu mới":"content-save-all",
				},
				callback=self.slec_hs
			)
		self.add_widget(self.Action_Button)
		
		self.getpath_Layout=MDFileManager(
			exit_manager=lambda x:self.getpath_Layout.close(),
			select_path=self.path,
			use_access=True,
			selector="folder",
			ext=[".Nonee"]
		)

		self.on_popup=ScrollView(
			size=(Window.width,Window.height),
			do_scroll_x=True,
			do_scroll_y=True
		)

		self.layout_themhs=Layout_Themhocsinh(self.app)

	def slec_hs(self,data):
		self.Action_Button.close_stack()
		if data.icon=="chart-bar":
			app.tab.transition.direction="left"
			app.tab.current="Chart"
			app.Chart.slec_tab()
		if data.icon=="account-plus":
			self.add_widget(self.on_popup)
			self.add_widget(self.layout_themhs)
		if data.icon=="content-save-all":
			self.app.open_loadfile(True)
		if data.icon=="content-save":
			self.app.save()
		if data.icon=="account-minus":
			if self.datatb.indexx == []:
				thongbao('Chọn học sinh muốn xóa')
				return
			for i in self.datatb.indexx:
				if i < len(self.datatb.row_data):self.app.xoahocsinh(False,self.datatb.row_data[i][0],self.datatb.row_data[i][1],self.datatb.row_data[i][2],self.datatb.row_data[i][3])
			self.app.datatables_rows_update()
		if data.icon=="database":
			data=[]
			for i in self.datatb.indexx:
				if i < len(self.datatb.row_data):data.append((self.datatb.row_data[i][0],self.datatb.row_data[i][1],self.datatb.row_data[i][2],self.datatb.row_data[i][3]))
			if data==[]:
				thongbao("Chọn học sinh muốn xuất dữ liệu")
				return
			path=self.getfilepath_Excel()
			self.data_xuatfile_excel=data
	def path(self,path):
		self._path_=path
		self.add_widget(self.app.Popup_on)
		self.add_widget(getfilepath_Excel())
		self.getpath_Layout.close()
	def getfilepath_Excel(self):
		if platform == 'android':
			self.getpath_Layout.show("/storage/emulated/0")
		else:
			self.getpath_Layout.show("/Users/Administrator")
	def xuatdulieu(self,filepath,datahs):
		wb = Workbook()
		ws_active=wb.active
		#Sheet chính ***********************************************************************************************************************
		ws_active.title="Bảng điểm tổng kết"
		ws_active.append(["Mã-HS","Họ và tên","NTNS","Lớp","Phái","ĐIỂM TRUNG BÌNH CÁC MÔN HK1"])
		tatcamonhoc=[]
		cot=6
		for data in datahs:
			hs = self.app.geths(data[0],data[1],data[2],data[3])
			if hs.diemhs != [] and hs.TBTCM(1) != 0 or hs.TBTCM(2) != 0:
				for mon in hs.diemhs:
					if mon.ten in tatcamonhoc or mon.dhk1==[]:
						pass
					else:
						ws_active[getcl(cot)+"2"].value=mon.ten
						cot+=1
						tatcamonhoc.append(mon.ten)
		if tatcamonhoc == []:
			thongbao('Học sinh đã chọn chưa có kết quả học tập')
			return
		ws_active.move_range(getcl(7)+"1:"+getcl(cot+10)+"1",rows=0,cols=cot-5)
		ws_active.merge_cells(getcl(6)+"1:"+getcl(cot-1)+"1")
		ws_active.merge_cells("A1:A2")
		ws_active.merge_cells("B1:B2")
		ws_active.merge_cells("C1:C2")
		ws_active.merge_cells("D1:D2")
		ws_active.merge_cells("E1:E2")
		ws_active.merge_cells(getcl(cot)+"1:"+getcl(cot)+"2")
		ws_active[getcl(cot)+"1"]="TB các môn HK1"
		hang=3
		for data in datahs:
			hs = self.app.geths(data[0],data[1],data[2],data[3])
			if hs.diemhs != [] and hs.TBTCM(1) != 0 or hs.TBTCM(2) != 0:
				ws_active["A"+str(hang)].value=hs.truong
				ws_active["B"+str(hang)].value=hs.ten
				ws_active["C"+str(hang)].value=hs.NTNS
				ws_active["D"+str(hang)].value=hs.lop
				ws_active["E"+str(hang)].value=hs._phai()[1]
				for i in range(6,cot):
					for mon in hs.diemhs:
						if mon.ten == ws_active[getcl(i)+"2"].value:
							ws_active[getcl(i)+str(hang)].value=mon.dtbmon(1)
				ws_active[getcl(cot)+str(hang)].value=hs.TBTCM(1)
				hang+=1
		cot+=1
		ws_active[getcl(cot)+"1"].value="Xếp loại"
		hang=3
		for data in datahs:
			hs = self.app.geths(data[0],data[1],data[2],data[3])
			if hs.diemhs != [] and hs.TBTCM(1) != 0 or hs.TBTCM(2) != 0:
				ws_active[getcl(cot)+str(hang)].value=hs.xeploai(1)
				hang+=1
		ws_active.merge_cells(getcl(cot)+"1:"+getcl(cot)+"2")
		tmp=False
		for data in datahs:
			hs = self.app.geths(data[0],data[1],data[2],data[3])
			if hs.diemhs != [] and hs.TBTCM(1) != 0 or hs.TBTCM(2) != 0:
				for mon in hs.diemhs:
					if mon.dhk2==[]:
						tmp=False
					else:
						for c in mon.dhk2:
							if c.diem!=[''] and c.diem!=[]:
								tmp=True
								break
						else:
							tmp=False
				if tmp:
					break
		oldhang=hang
		if tmp:
			tatcamonhoc=[]
			cot+=1
			oldcot=cot
			ws_active[getcl(cot)+"1"].value="ĐIỂM TRUNG BÌNH CÁC MÔN HK2"
			ws_active[getcl(cot+1)+"1"].value="TB các môn HK2"
			for data in datahs:
				hs = self.app.geths(data[0],data[1],data[2],data[3])
				if hs.diemhs != [] and hs.TBTCM(1) != 0 or hs.TBTCM(2) != 0:
					for mon in hs.diemhs:
						if mon.ten in tatcamonhoc or mon.dhk2==[]:
							pass
						else:
							ws_active[getcl(cot)+"2"].value=mon.ten
							cot+=1
							tatcamonhoc.append(mon.ten)
			hang=3
			for data in datahs:
				hs = self.app.geths(data[0],data[1],data[2],data[3])
				if hs.diemhs != [] and hs.TBTCM(1) != 0 or hs.TBTCM(2) != 0:
					ws_active["A"+str(hang)].value=hs.truong
					ws_active["B"+str(hang)].value=hs.ten
					ws_active["C"+str(hang)].value=hs.NTNS
					ws_active["D"+str(hang)].value=hs.lop
					ws_active["E"+str(hang)].value=hs._phai()[1]
					for i in range(oldcot,cot):
						for mon in hs.diemhs:
							if mon.ten == ws_active[getcl(i)+"2"].value:
								ws_active[getcl(i)+str(hang)].value=mon.dtbmon(2)
					ws_active[getcl(cot)+str(hang)].value=hs.TBTCM(2)
					hang+=1
			ws_active.move_range(getcl(oldcot+1)+"1:"+getcl(cot+10)+"1",rows=0,cols=cot-oldcot-1)
			ws_active.merge_cells(getcl(oldcot)+"1:"+getcl(cot-1)+"1")
			ws_active.merge_cells(getcl(cot)+"1:"+getcl(cot)+"2")
			cot+=1
			ws_active[getcl(cot)+"1"].value="Xếp loại"
			hang=3
			for data in datahs:
				hs = self.app.geths(data[0],data[1],data[2],data[3])
				if hs.diemhs != [] and hs.TBTCM(1) != 0 or hs.TBTCM(2) != 0:
					ws_active[getcl(cot)+str(hang)].value=hs.xeploai(2)
					hang+=1
			ws_active.merge_cells(getcl(cot)+"1:"+getcl(cot)+"2")
		if oldhang > hang:
			hang = oldhang
		for hangg in range(1,hang):
			for cott in range(1,cot+1):
				ws_active[getcl(cott)+str(hangg)].alignment = Alignment(horizontal='center',vertical = 'center')
				if hangg<3:ws_active[getcl(cott)+str(hangg)].fill = PatternFill(fgColor='b7b7b7', fill_type='solid')
				self.set_border(ws_active,getcl(cott)+str(hangg))
				if hangg<3 or cott<6:
					ws_active[getcl(cott)+str(hangg)].font=Font(bold=True)
				if ws_active[getcl(cott)+str(hangg)].value != None and ws_active.column_dimensions[getcl(cott)].width < len(str(ws_active[getcl(cott)+str(hangg)].value))+4:
					ws_active.column_dimensions[getcl(cott)].width=len(str(ws_active[getcl(cott)+str(hangg)].value))+4
		#Các sheet nhỏ*********************************************************************************************************************
		ws = []
		for data in datahs:
			hs = self.app.geths(data[0],data[1],data[2],data[3])
			if hs.diemhs != [] and hs.TBTCM(1) != 0 or hs.TBTCM(2) != 0:
				for mon in hs.diemhs:
					if mon.ten in wb.sheetnames:
						pass
					else:
						ws.append(wb.create_sheet(mon.ten))
		for sheet in ws:
			tonghs=[]
			sheet.append(["Mã-HS","Họ và tên","NTNS","Lớp","Phái","Học kì 1"])
			tatcacotdiem=[]
			cot=6
			for data in datahs:
				hs = self.app.geths(data[0],data[1],data[2],data[3])
				if hs.diemhs != [] and hs.TBTCM(1) != 0 or hs.TBTCM(2) != 0:
					for mon in hs.diemhs:
						if mon.ten == sheet.title:
							if hs.ten in tonghs:
								pass
							else:
								tonghs.append(hs.ten)
							cotd=mon.dhk1.copy()
							cotd.reverse()
							for cotdiem in cotd:
								if cotdiem.ten in tatcacotdiem or cotdiem.diem==[]:
									pass
								else:
									sheet[getcl(cot)+"2"].value=cotdiem.ten
									cot+=1
									tatcacotdiem.append(cotdiem.ten)
			sheet.move_range(getcl(7)+"1:"+getcl(cot+10)+"1",rows=0,cols=cot-5)
			if cot != 6:sheet.merge_cells(getcl(6)+"1:"+getcl(cot-1)+"1")
			sheet.merge_cells(getcl(cot)+"1:"+getcl(cot)+"2")
			sheet.merge_cells("A1:A2")
			sheet.merge_cells("B1:B2")
			sheet.merge_cells("C1:C2")
			sheet.merge_cells("D1:D2")
			sheet.merge_cells("E1:E2")
			sheet[getcl(cot)+"1"]="TB Môn học kì1"
			hang=3
			for data in datahs:
				hs = self.app.geths(data[0],data[1],data[2],data[3])
				if hs.diemhs != [] and hs.TBTCM(1) != 0 or hs.TBTCM(2) != 0:
					oldhang=hang
					for mon in hs.diemhs:
						if mon.ten == sheet.title:
							sheet["A"+str(hang)].value=hs.truong
							sheet["B"+str(hang)].value=hs.ten
							sheet["C"+str(hang)].value=hs.NTNS
							sheet["D"+str(hang)].value=hs.lop
							sheet["E"+str(hang)].value=hs._phai()[1]
							for cotdiem in mon.dhk1:
								for i in range(6,cot):
									if cotdiem.ten == sheet[getcl(i)+"2"].value:
										hangdiem=oldhang
										for diemm in cotdiem.diem:
											if diemm == "":diemm=0
											sheet[getcl(i)+str(hangdiem)].value=float(diemm)
											hangdiem+=1
											if hang < hangdiem:
												hang=hangdiem
										if cotdiem.diem==[]:
											sheet[getcl(i)+str(hangdiem)].value=0
											hangdiem+=1
											if hang < hangdiem:
												hang=hangdiem
							for cotdiem in mon.dhk2:
								if hang < oldhang+len(cotdiem.diem):
									hang=oldhang+len(cotdiem.diem)
							sheet[getcl(cot)+str(oldhang)].value=mon.dtbmon(1)

							for cotdiem in mon.dhk1:
								for i in range(6,cot):
									if cotdiem.ten == sheet[getcl(i)+"2"].value:
										for j in range(3,hang):
											if sheet[getcl(i)+str(j)].value != None:
												old=[i,j]
										else:
											sheet.merge_cells(getcl(old[0])+str(old[1])+":"+getcl(i)+str(j))

			tmp=False
			for data in datahs:
				hs = self.app.geths(data[0],data[1],data[2],data[3])
				if hs.diemhs != [] and hs.TBTCM(1) != 0 or hs.TBTCM(2) != 0:
					for mon in hs.diemhs:
						if mon.ten == sheet.title:
							if mon.dhk2==[]:
								tmp=False
							else:
								for c in mon.dhk2:
									if c.diem!=[''] and c.diem!=[]:
										tmp=True
										break
								else:
									tmp=False
					if tmp:
						break
			oldhang=hang
			cot+=1
			oldcot=cot
			if tmp:
				tatcacotdiem=[]
				sheet[getcl(cot)+"1"].value="Học kì2"
				sheet[getcl(cot+1)+"1"].value="TB Môn học kì2"
				for data in datahs:
					hs = self.app.geths(data[0],data[1],data[2],data[3])
					if hs.diemhs != [] and hs.TBTCM(1) != 0 or hs.TBTCM(2) != 0:
						for mon in hs.diemhs:
							if mon.ten == sheet.title:
								cotd=mon.dhk2.copy()
								cotd.reverse()
								for cotdiem in cotd:
									if cotdiem.ten in tatcacotdiem or cotdiem.diem==[]:
										pass
									else:
										sheet[getcl(cot)+"2"].value=cotdiem.ten
										cot+=1
										tatcacotdiem.append(cotdiem.ten)
				hang=3
				for data in datahs:
					hs = self.app.geths(data[0],data[1],data[2],data[3])
					if hs.diemhs != [] and hs.TBTCM(1) != 0 or hs.TBTCM(2) != 0:
						for mon in hs.diemhs:
							if mon.ten == sheet.title:
								old=hang
								for cotdiem in mon.dhk2:
									for i in range(oldcot,cot):
										if cotdiem.ten == sheet[getcl(i)+"2"].value:
											hangdiem=old
											for diemm in cotdiem.diem:
												if diemm == "":diemm=0
												sheet[getcl(i)+str(hangdiem)].value=float(diemm)
												hangdiem+=1
												if hang < hangdiem:
													hang=hangdiem
											if cotdiem.diem==[]:
												sheet[getcl(i)+str(hangdiem)].value=0
												hangdiem+=1
												if hang < hangdiem:
													hang=hangdiem
								for cotdiem in mon.dhk1:
									if hang < old+len(cotdiem.diem):
										hang=old+len(cotdiem.diem)
								sheet[getcl(cot)+str(old)].value=mon.dtbmon(2)

								for cotdiem in mon.dhk2:
									for i in range(oldcot,cot):
										if cotdiem.ten == sheet[getcl(i)+"2"].value:
											for j in range(3,hang):
												if sheet[getcl(i)+str(j)].value != None:
													old=[i,j]
											else:
												sheet.merge_cells(getcl(old[0])+str(old[1])+":"+getcl(i)+str(j))
				sheet.move_range(getcl(oldcot+1)+"1:"+getcl(cot+10)+"1",rows=0,cols=cot-oldcot-1)
				sheet.merge_cells(getcl(oldcot)+"1:"+getcl(cot-1)+"1")
				sheet.merge_cells(getcl(cot)+"1:"+getcl(cot)+"2")
			else:
				cot-=1
				oldcot=cot
			if oldhang > hang:
				hang=oldhang
			j=3
			for i in range(3,hang):
				if sheet["B"+str(i)].value in tonghs and j != i:
					sheet.merge_cells("A"+str(j)+":A"+str(i-1))
					sheet.merge_cells("B"+str(j)+":B"+str(i-1))
					sheet.merge_cells("C"+str(j)+":C"+str(i-1))
					sheet.merge_cells("D"+str(j)+":D"+str(i-1))
					sheet.merge_cells("E"+str(j)+":E"+str(i-1))
					if oldcot != cot:
						sheet.merge_cells(getcl(oldcot-1)+str(j)+":"+getcl(oldcot-1)+str(i-1))
					sheet.merge_cells(getcl(cot)+str(j)+":"+getcl(cot)+str(i-1))
					j=i
			else:
				sheet.merge_cells("A"+str(j)+":A"+str(i))
				sheet.merge_cells("B"+str(j)+":B"+str(i))
				sheet.merge_cells("C"+str(j)+":C"+str(i))
				sheet.merge_cells("D"+str(j)+":D"+str(i))
				sheet.merge_cells("E"+str(j)+":E"+str(i))
				if oldcot != cot:
					sheet.merge_cells(getcl(oldcot-1)+str(j)+":"+getcl(oldcot-1)+str(i))
				sheet.merge_cells(getcl(cot)+str(j)+":"+getcl(cot)+str(i))
			for cott in range(1,cot+1):
				sheet.column_dimensions[getcl(cot)].width=1
				for hangg in range(1,hang):
					sheet[getcl(cott)+str(hangg)].alignment = Alignment(horizontal='center',vertical = 'center')
					if hangg<3:sheet[getcl(cott)+str(hangg)].fill = PatternFill(fgColor='b7b7b7', fill_type='solid')
					self.set_border(sheet,getcl(cott)+str(hangg))
					if hangg<3 or cott<6:sheet[getcl(cott)+str(hangg)].font=Font(bold=True)
					if sheet[getcl(cott)+str(hangg)].value != None and sheet.column_dimensions[getcl(cott)].width < len(str(sheet[getcl(cott)+str(hangg)].value))+2:
						sheet.column_dimensions[getcl(cott)].width=len(str(sheet[getcl(cott)+str(hangg)].value))+2
		try:
			wb.save(filepath+".xlsx")
		except:
			thongbao("Lỗi: Hãy xóa cửa sổ excel đang hoạt động và thử lại.")
			return
		thongbao(f"Đã xuất dữ liệu thành công ---> {filepath}.xlsx")
	def set_border(self,ws,cell_range):
	    thin = Side(border_style="thin", color="000000")
	    ws[cell_range].border = Border(top=thin, left=thin, right=thin, bottom=thin)
class Layout_Themhocsinh(GridLayout):
	def __init__(self,*args,**kwargn):
		super(Layout_Themhocsinh,self).__init__(**kwargn)
		self.app=args[0]
		self.ALL_TextField=[]

		self.cols=1
		self.pos_hint={"center_x":.5,"center_y":.5}
		self.size_hint=(None,None)
		self.height=400
		self.width=700
		with self.canvas.before:
			Color(rgba=(62/256, 112/256, 128/256, 1))
			self.rect=RoundedRectangle(radius=[(40.0, 40.0), (40.0, 40.0), (40.0, 40.0), (40.0, 40.0)])
		self.bind(pos=self.update_rect,size=self.update_rect)
		
		self.Xong_button=MDFillRoundFlatButton(text="Thêm")
		self.Huy_button=MDFillRoundFlatButton(text="Hủy")
		self.Xong_button.huy=self.Huy_button
		self.Huy_button.xong=self.Xong_button
		self.Xong_button.bind(on_press=self.Xong_)
		self.Huy_button.bind(on_press=self.Huy_)
#-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*Tạo 2 mãng giao diện
		self.Thongtinhs_layout=GridLayout(rows=1,padding=10,size_hint_y=None,height=100)
		self.add_widget(self.Thongtinhs_layout)
		self.ngay=MDTextField(
				font_name="viet",
				hint_text="Ngày",
				text_color=(0/256, 1/256, 73/256, 1),
				halign="center",
				error=True,
				helper_text='...\nNgày sinh\nkhông thể bỏ trống'
			)
		self.ALL_TextField.append(self.ngay)
		self.ngay.bind(text=self.Textfield_NTNS)
		self.ngay.bind(focus=self.Textfield_NTNS)
		self.thang=MDTextField(
				font_name="viet",
				hint_text="Tháng",
				text_color=(0/256, 1/256, 73/256, 1),
				halign="center",
				error=True,
				helper_text='...\nTháng sinh\nkhông thể bỏ trống'
			)
		self.ALL_TextField.append(self.thang)
		self.thang.bind(text=self.Textfield_NTNS)
		self.thang.bind(focus=self.Textfield_NTNS)
		self.nam=MDTextField(
				font_name="viet",
				hint_text="Năm",
				text_color=(0/256, 1/256, 73/256, 1),
				halign="center",
				error=True,
				helper_text='...\nNăm sinh\nkhông thể bỏ trống'
			)
		self.ALL_TextField.append(self.nam)
		self.nam.bind(text=self.Textfield_NTNS)
		self.nam.bind(focus=self.Textfield_NTNS)
		self.layout_NTNS_picker=GridLayout(rows=1,pos_hint={'center_x':.5,'center_y':.5},size_hint=(None,None),width=500)
		tmp=Screen(size_hint_x=1,size_hint_y=1)
		tmp.add_widget(self.layout_NTNS_picker)
		self.layout_NTNS_picker.add_widget(self.ngay)
		self.layout_NTNS_picker.add_widget(MDLabel(text="/",size_hint_x=None,width=50,halign="center"))
		self.layout_NTNS_picker.add_widget(self.thang)
		self.layout_NTNS_picker.add_widget(MDLabel(text="/",size_hint_x=None,width=50,halign="center"))
		self.layout_NTNS_picker.add_widget(self.nam)
		
		self.layout_NTNS=GridLayout(cols=1,size_hint_y=None,height=150,padding=10)
		self.layout_NTNS.add_widget(MDLabel(text='Ngày tháng năm sinh: ',size_hint=(None,None),center_x=.5,width=400))
		self.layout_NTNS.add_widget(tmp)
		
		self.add_widget(self.layout_NTNS)
#-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*Thiết lập giao diện thông tin hs
		self.Truong_TextField=MDTextField(
				hint_text="Mã-HS",
				font_name="viet",
				text_color=(0/256, 1/256, 73/256, 1),
				error=True,
				halign="center",
				helper_text='Mã-HS không thể bỏ trống'
			)
		self.ALL_TextField.append(self.Truong_TextField)
		self.Truong_TextField.bind(text=self.on_edit)
		self.Truong_TextField.bind(focus=self.Textfield_Checkerror)
		self.Truong_TextField.bind(text=self.Textfield_Checkerror)
		self.Lop_TextField=MDTextField(
				hint_text="Tên lớp",
				font_name="viet",
				text_color=(0/256, 1/256, 73/256, 1),
				error=True,
				size_hint_x=None,
				halign="center",
				width=200,
				helper_text='Tên lớp không thể bỏ trống'
			)
		self.ALL_TextField.append(self.Lop_TextField)
		self.Lop_TextField.bind(text=self.on_edit)
		self.Lop_TextField.bind(focus=self.Textfield_Checkerror)
		self.Lop_TextField.bind(text=self.Textfield_Checkerror)
		self.Hs_TextField=MDTextField(
				hint_text="Tên học sinh",
				font_name="viet",
				text_color=(0/256, 1/256, 73/256, 1),
				error=True,
				halign="center",
				helper_text='Tên học sinh không thể bỏ trống'
			)

		self.ALL_TextField.append(self.Hs_TextField)
		self.Hs_TextField.bind(text=self.Textfield_Checkerror)
		self.Hs_TextField.bind(focus=self.Textfield_Checkerror)
		self.Thongtinhs_layout.add_widget(self.Truong_TextField)
		self.Thongtinhs_layout.add_widget(self.Lop_TextField)
		self.Thongtinhs_layout.add_widget(self.Hs_TextField)
		self.menu=MDDropdownMenu(
					caller=self.Lop_TextField,
					max_height=500,
					width_mult=4,
					position="bottom"
				)
				
		
		self.phai=MDIconButton(icon='human-male')
		self.phai.bind(on_release=self.change_sex)
		self.phai.pos=(self.pos[0]+self.width/2-self.phai.width/2,self.pos[1])

		tmp=Screen()
		tmp.add_widget(self.phai)
		tmp.add_widget(self.Xong_button)
		tmp.add_widget(self.Huy_button)
		self.add_widget(tmp)
		self.Xong_button.pos=(self.pos[0]+self.width-136,self.pos[1])
		self.xongbutton=False
		self.settext=True
#-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*Funtion - Giao diện thông tin HS
	def change_sex(self,*args):
		if self.phai.icon=='human-male':
			self.phai.icon='human-female'
		else:
			self.phai.icon='human-male'
	def Textfield_Checkerror(self,*args):
		if args[0].text=="" and args[0].focus:
			args[0].error=True
			if args[0].hint_text=='Tên học sinh':args[0].helper_text='Tên học sinh không thể bỏ trống'
			elif args[0].hint_text=='Tên lớp':args[0].helper_text='Tên lớp không thể bỏ trống'
			elif args[0].hint_text=='Mã-HS':args[0].helper_text='Mã-HS không thể bỏ trống'
			args[0]._anim_current_line_color(args[0].error_color)
		elif args[0].text=="":
			args[0].helper_text=''
			args[0]._anim_current_line_color(args[0].error_color)
		else:
			args[0].error=False
			args[0].helper_text=''
			args[0]._anim_current_line_color(args[0].theme_cls.disabled_hint_text_color)
	def Textfield_NTNS(self,*args):
		if args[0].text=="" and args[0].focus:
			args[0].error=True
			if args[0].hint_text=='Ngày':args[0].helper_text='...\nNgày sinh\nkhông thể bỏ trống'
			elif args[0].hint_text=='Tháng':args[0].helper_text='...\nTháng sinh\nkhông thể bỏ trống'
			elif args[0].hint_text=='Năm':args[0].helper_text='...\nNăm sinh\nkhông thể bỏ trống'
			args[0]._anim_current_line_color(args[0].error_color)
		elif args[0].focus and args[0].hint_text=='Ngày' or args[0].hint_text=='Tháng' and args[0].focus:
			args[0].error=True
			try:
				if args[0].hint_text=='Ngày' and int(args[0].text)>31 or int(args[0].text)<=0 and args[0].hint_text=='Ngày':
					args[0]._anim_current_line_color(args[0].error_color)
					args[0].helper_text='Lỗi định dạng ngày sinh'
				elif args[0].hint_text=='Tháng' and int(args[0].text)>12 or int(args[0].text)<=0 and args[0].hint_text=='Tháng':
					args[0]._anim_current_line_color(args[0].error_color)
					args[0].helper_text='Lỗi định dạng Tháng sinh'
				else:
					args[0].error=False
					args[0].helper_text=''
					args[0]._anim_current_line_color(args[0].theme_cls.disabled_hint_text_color)
			except:
				if args[0].hint_text=='Ngày':
					args[0]._anim_current_line_color(args[0].error_color)
					args[0].helper_text='Lỗi định dạng ngày sinh'
				elif args[0].hint_text=='Tháng':
					args[0]._anim_current_line_color(args[0].error_color)
					args[0].helper_text='Lỗi định dạng Tháng sinh'
				else:
					args[0].error=False
					args[0].helper_text=''
					args[0]._anim_current_line_color(args[0].theme_cls.disabled_hint_text_color)
		elif args[0].hint_text=='Ngày' or args[0].hint_text=='Tháng':
			try:
				if args[0].hint_text=='Ngày' and int(args[0].text)>31 or args[0].hint_text=='Ngày' and int(args[0].text)<=0:
					args[0].helper_text=''
					args[0]._anim_current_line_color(args[0].error_color)
				elif args[0].hint_text=='Tháng' and int(args[0].text)>12 or args[0].hint_text=='Tháng' and int(args[0].text)<=0:
					args[0].helper_text=''
					args[0]._anim_current_line_color(args[0].error_color)
			except:
				if args[0].hint_text=='Ngày':
					args[0].helper_text=''
					args[0]._anim_current_line_color(args[0].error_color)
				elif args[0].hint_text=='Tháng':
					args[0].helper_text=''
					args[0]._anim_current_line_color(args[0].error_color)
		elif args[0].focus and args[0].hint_text=='Năm':
			args[0].error=True
			try:
				nam = int(args[0].text)
				args[0].error=False
				args[0].helper_text=''
				args[0]._anim_current_line_color(args[0].theme_cls.disabled_hint_text_color)
			except:
				args[0]._anim_current_line_color(args[0].error_color)
				args[0].helper_text='Lỗi định dạng Năm sinh'
		elif args[0].text=="":
			args[0].helper_text=''
			args[0]._anim_current_line_color(args[0].error_color)
		else:
			args[0].error=False
			args[0].helper_text=''
			args[0]._anim_current_line_color(args[0].theme_cls.disabled_hint_text_color)
	def slec_truong(self,*args):
		if args[1]=="T":
			self.settext=False
			self.Truong_TextField.text=args[0]
		if args[1]=="L":
			self.settext=False
			self.Lop_TextField.text=args[0]
		self.menu.dismiss()
	def on_edit(self,*args):
		if self.xongbutton:
			return
		if self.settext and args[1]!='':
			self.menu.items=[]
			self.menu.dismiss()
			if args[0].hint_text=="Tên lớp":
				all_lop=[]
				for tr in app.truong:
					for lop in tr.lop:
						if args[1].lower() in lop.ten.lower() or args[1].upper() in lop.ten.upper() or args[1] in lop.ten:
							if lop.ten not in all_lop:
								self.menu.items.append(
										{
											"text":lop.ten,
											"viewclass": "OneLineListItem",
											"on_release":lambda x=lop.ten:self.slec_truong(x,"L"),
										}
									)
								all_lop.append(lop.ten)
			if self.menu.items!=[]:self.menu.open()
		else:
			self.menu.dismiss()
			self.settext=True
#-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*Funtion - Giao diện Button
	def Xong_(self,*args):
		self.xongbutton=True
		for i in self.ALL_TextField:
			if i.error:
				thongbao(f"Lỗi không thể thêm học sinh: {i.hint_text}")
				return
		for tr in self.app.truong:
			if tr.ten == self.Truong_TextField.text:
				thongbao("Mã học sinh không thể trùng với học sinh khác")
				return
			for lop in tr.lop:
				for hs in lop.hocsinh:
					if hs.ten.lower() == self.Hs_TextField.text.lower() and hs.NTNS == str(self.ngay.text)+'/'+str(self.thang.text)+'/'+str(self.nam.text):
						thongbao("Học sinh này đã có trong dữ liệu")
						return
		self.app.truong.append(Ctruong(self.Truong_TextField.text))
		for tr in self.app.truong:
			for lop in tr.lop:
				if lop.ten.lower() == self.Lop_TextField.text.lower():
					self.Lop_TextField.text=lop.ten
		self.app.truong[-1].lop.append(Clop(self.Lop_TextField.text))
		for tr in self.app.truong:
			if tr.ten.lower() == self.Truong_TextField.text.lower():
				for lop in tr.lop:
					if lop.ten.lower() == self.Lop_TextField.text.lower():
						lop.hocsinh.append(Chocsinh(str(self.Hs_TextField.text),str(self.ngay.text)+'/'+str(self.thang.text)+'/'+str(self.nam.text)))
						lop.hocsinh[-1].truong=tr.ten
						lop.hocsinh[-1].lop=lop.ten
						if self.phai.icon=='human-male':
							lop.hocsinh[-1].phai=True
						else:
							lop.hocsinh[-1].phải=False
		self.app.datatables_rows_update()
		self.app.screen.remove_widget(self.app.screen.layout_themhs)
		self.app.screen.remove_widget(self.app.screen.on_popup)
		for i in self.ALL_TextField:
			i.text=""
		self.xongbutton=False
	def Huy_(self,*args):
		self.app.screen.remove_widget(self.app.screen.layout_themhs)
		self.app.screen.remove_widget(self.app.screen.on_popup)
		self.app.screen.remove_widget(args[0].xong)
		self.app.screen.remove_widget(args[0])
		for i in self.ALL_TextField:
			i.text=""
	def update_rect(self,rect,a):
		rect.rect.pos=rect.pos
		rect.rect.size=rect.size
class Myapp(MDApp):
	def build(self):
		Window.bind(size=self.resize)
		self.theme_cls.theme_style = "Light"
		self.theme_cls.primary_palette = "BlueGray"
		self.tab=ScreenManager()
		self.screen = Layout_Hocsinh(self)
		self.check_press=False
		self.screen.name="Main"
		self.screen.bind(on_touch_down=self.touch_down)
		self.screen.bind(on_touch_up=self.touch_up)
		self.item_menubar=BoxLayout(spacing=8)
		self.tabdiem=diemlayout(self)
		self.tabdiem.menubar.left_action_items=[["backspace", lambda x: self.move_tab("Main")]]
		

		self.Toolbar=MDToolbar(
				title="Quản lý sinh viên",
				pos_hint={'top':1},
				type="top",
				elevation=10,
		)
		self.Toolbar.left_action_items=[["school", lambda x: self.nav_drawer.set_state("open")]]
		self.Toolbar.right_action_items=[["account-search",self.lochs]]


		self.nav_drawer=MDNavigationDrawer()
		self.item_menubar.orientation='vertical'
		self.item_menubar.add_widget(
			MDToolbar(
				title="Menu",
			)
		)
		self.item_menubar.Loadfile=OneLineIconListItem(
			text='Nhập dữ liệu',
			size_hint_y=None
		)
		self.item_menubar.Loadfile.add_widget(IconLeftWidget(icon="note"))
		self.item_menubar.Loadfile.bind(on_release=self.Load_file)

		self.item_menubar.LoadfileExcel=OneLineIconListItem(
			text='Nhập dữ liệu từ Excel',
			size_hint_y=None
		)
		self.item_menubar.LoadfileExcel.bind(on_release=self.Load_file_Excel)
		self.item_menubar.LoadfileExcel.add_widget(IconLeftWidget(icon="database-plus"))
		self.item_menubar.add_widget(
			self.item_menubar.Loadfile
		)
		self.item_menubar.add_widget(
			self.item_menubar.LoadfileExcel
		)
		self.item_menubar.add_widget(ScrollView())
		self.nav_drawer.add_widget(self.item_menubar)





		self.truong=[]
		self.monhoc=[]
		self.loi=[]


		self.screen.add_widget(self.Toolbar)
		self.screen.add_widget(self.nav_drawer)


		self.Loadfile_Layout=MDFileManager(
			exit_manager=lambda x:self.Loadfile_Layout.close(),
			select_path=self.select_path,
			selector="file",
			use_access=True,
			ext=[".LTP"]
		)
		self.Savefile_Layout=MDFileManager(
			exit_manager=lambda x:self.Savefile_Layout.close(),
			select_path=self.select_pathsave,
			use_access=True,
			ext=[".LTP"]
		)
		self.Loadfile_Excel_Layout=MDFileManager(
			exit_manager=lambda x:self.Loadfile_Excel_Layout.close(),
			select_path=self.select_path_excel_loadfile,
			selector="file",
			use_access=True,
			ext=[".xlsx"]
		)

		self.Timeload_layout=layout_timeload()
		self.tab.add_widget(self.screen)
		self.tab.add_widget(self.tabdiem)
		self.tab.add_widget(self.Timeload_layout)
		self.tab.transition.bind(on_complete=self.TimeLoad_complete)
		self.tab.transition.duration=.4


		self.Chart=Layout_Chart()
		self.tab.add_widget(self.Chart)


		self.Popup_on=ScrollView(
			size=(Window.width,Window.height),
			do_scroll_x=True,
			do_scroll_y=True
		)
		self.layout_thietlapmon=layout_thietlapmon(self)
		self.tabdiem.menubar.right_action_items=[["book",self.open_thietlapmon]]

		self.thietlapmon_xong=MDFillRoundFlatButton(
						text="Xong",
						on_release=self._thietlapmon_xong,
						font_size=sp(20),
						pos=(100,100)
					)
		self.thietlapmon_xong.heso_TextField=[]
		self.thietlapmon_xong.mon_TextField=[]
		self.thietlapmon_xong.cotdiem_TextField=[]
		self.save_tab="main"
		
		
		try:
			if platform == "android":
				from android.permissions import request_permissions, Permission
				request_permissions([Permission.READ_EXTERNAL_STORAGE, Permission.WRITE_EXTERNAL_STORAGE])
		except:
			pass
		return self.tab



	def lochs(self,*args):
		self.menu_loc=MDDropdownMenu(
				caller=args[0],
				max_height=500,
				items=[
						{
							"text":"Tìm học sinh",
							"viewclass": "OneLineListItem",
							"on_release":lambda x=1:self.loc(x),
						},
						{
							"text":"Lọc học sinh theo phái",
							"viewclass": "OneLineListItem",
							"on_release":lambda x=4:self.loc(x),
						},
						{
							"text":"Tìm lớp học",
							"viewclass": "OneLineListItem",
							"on_release":lambda x=2:self.loc(x),
						},
						{
							"text":"Lọc học sinh theo điểm trung bình môn",
							"viewclass": "OneLineListItem",
							"on_release":lambda x=3:self.loc(x),
						},
					],
				width_mult=6
			)
		self.menu_loc.open()
	def loc(self,type,*args):
		self.menu_loc.dismiss()
		if type==1:
			self.menu_loc=MDDropdownMenu(
					caller=self.menu_loc.caller,
					max_height=500,
					items=[
							{
								"text":"Tìm bằng mã học sinh",
								"viewclass": "OneLineListItem",
								"on_release":lambda x=1.1:self.loc(x),
							},
							{
								"text":"Tìm bằng tên",
								"viewclass": "OneLineListItem",
								"on_release":lambda x=1.2:self.loc(x),
							},
						],
					width_mult=5
				)
			self.menu_loc.open()
		elif type==1.1:
			layout=layout_timhocsinh_MHS()
			self.screen.add_widget(self.Popup_on)
			self.screen.add_widget(layout)
		elif type==1.2:
			layout=layout_timhocsinh_ten()
			self.screen.add_widget(self.Popup_on)
			self.screen.add_widget(layout)
		elif type==2:
			self.menu_loc=MDDropdownMenu(
					caller=self.menu_loc.caller,
					max_height=500,
					width_mult=4
				)
			all_lop=[]
			for tr in self.truong:
				for lop in tr.lop:
					if lop.ten not in all_lop:
						self.menu_loc.items.append(
								{
									"text":lop.ten,
									"viewclass": "OneLineListItem",
									"on_release":lambda x=lop.ten:self.loc(2.1,x),
								}
							)
						all_lop.append(lop.ten)
			if all_lop==[]:
				thongbao("không có lớp học nào trong cơ sở dữ liệu")
				return
			self.menu_loc.open()
		elif type==2.1:
			listt=[]
			for tr in self.truong:
				for lop in tr.lop:
					if lop.ten == args[0]:
						for hs in lop.hocsinh:
							row=[]
							row.append(tr.ten)
							row.append(lop.ten)
							row.append(hs.ten)
							row.append(hs.NTNS)
							row.append(hs._phai())
							listt.append(tuple(row))
			self.screen.datatb.row_data=listt
			self.Toolbar.right_action_items=[["close",self.datatables_rows_update]]
		elif type==3:
			layout=layout_lochs()
			self.screen.add_widget(self.Popup_on)
			self.screen.add_widget(layout)
		elif type==4:
			self.menu_loc=MDDropdownMenu(
					caller=self.menu_loc.caller,
					max_height=500,
					items=[
							{
								"text":"Nam",
								"viewclass": "OneLineListItem",
								"on_release":lambda x=4.1:self.loc(x),
							},
							{
								"text":"Nữ",
								"viewclass": "OneLineListItem",
								"on_release":lambda x=4.2:self.loc(x),
							},
						],
					width_mult=5
				)
			self.menu_loc.open()
		elif type==4.1:
			listt=[]
			for tr in self.truong:
				for lop in tr.lop:
					for hs in lop.hocsinh:
						if hs.phai:
							row=[]
							row.append(tr.ten)
							row.append(lop.ten)
							row.append(hs.ten)
							row.append(hs.NTNS)
							row.append(hs._phai())
							listt.append(tuple(row))
			if listt==[]:
				thongbao('Không có học sinh nam trong cơ sở dữ liệu')
				return
			self.screen.datatb.row_data=listt
			self.Toolbar.right_action_items=[["close",self.datatables_rows_update]]
		elif type==4.2:
			listt=[]
			for tr in self.truong:
				for lop in tr.lop:
					for hs in lop.hocsinh:
						if not hs.phai:
							row=[]
							row.append(tr.ten)
							row.append(lop.ten)
							row.append(hs.ten)
							row.append(hs.NTNS)
							row.append(hs._phai())
							listt.append(tuple(row))
			if listt==[]:
				thongbao('Không có học sinh nữ trong cơ sở dữ liệu')
				return
			self.screen.datatb.row_data=listt
			self.Toolbar.right_action_items=[["close",self.datatables_rows_update]]
	#-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-
	def resize(self,window,size):
		self.Chart.MDTabs.height=Window.height-self.Chart.menubar.height
		self.tabdiem.View.width=size[0]
		self.tabdiem.View2.width=size[0]
	def _thietlapmon_xong(self,*args):
		for i in self.thietlapmon_xong.mon_TextField:
			if i.error:
				thongbao(f"Lỗi: Môn học-{i.error_text}")
				return
		for i in self.thietlapmon_xong.cotdiem_TextField:
			if i.error:
				thongbao(f"Lỗi: Cột điểm-{i.error_text}")
				return
		for i in self.thietlapmon_xong.heso_TextField:
			if i.error:
				thongbao(f"Lỗi: Hệ số-{i.error_text}")
				return
		self.tabdiem.remove_widget(self.layout_thietlapmon)
		self.tabdiem.remove_widget(self.thietlapmon_xong)
		self.tabdiem.remove_widget(self.Popup_on)
	def open_thietlapmon(self,*args):
		self.tabdiem.add_widget(self.Popup_on)
		self.tabdiem.add_widget(self.layout_thietlapmon)
		self.tabdiem.add_widget(self.thietlapmon_xong)
		self.layout_thietlapmon.open()
	#Trang Điểm#####################################################################################################################################################
	def TimeLoad_complete(self,*args):
		if self.tab.current=="TimeLoad":
			self.tab.transition.direction="left"
			self.tabdiem.tenhs.text=f"Tên học sinh: {self.data_hocsinh[2]}"
			load=True
			for mon in self.geths(self.data_hocsinh[0],self.data_hocsinh[1],self.data_hocsinh[2],self.data_hocsinh[3]).diemhs:
				self.themmon(mon.ten)
				for cotdiem in mon.dhk1:
					load=False
					for diem in cotdiem.diem:
						try:
							d=float(diem)
							self.themdiem(1,mon.ten,cotdiem.ten,diem)
						except:
							pass
				for cotdiem in mon.dhk2:
					load=False
					for diem in cotdiem.diem:
						try:
							d=float(diem)
							self.themdiem(2,mon.ten,cotdiem.ten,diem)
						except:
							pass
			if load:
				self.themmon_button=MDIconButton(icon="loupe")
				self.themmon_button.bind(on_press=self.menu_open)
				self.tabdiem.View.add_widget(self.themmon_button)
				self.themmon_button2=MDIconButton(icon="loupe")
				self.themmon_button2.bind(on_press=self.menu_open)
				self.tabdiem.View2.add_widget(self.themmon_button2)
			self.tab.current="tabdiem"
	def menu_open(self,*args):
		self.list=[]
		mondaco=[]
		for View in self.tabdiem.View.children[1:]:
			if len(View.children[1].children[0].children)==1:
				mondaco.append(View.children[1].children[0].children[0].text[9:])
			else:
				mondaco.append(View.children[1].children[0].children[1].text[9:])
		for i in self.monhoc:
			if i.ten not in mondaco:
				self.list.append(
						{
							"text":i.ten,
							"viewclass": "OneLineListItem",
							"on_release":lambda x=i.ten:self.press(x),
						}
					)
		self.menu=MDDropdownMenu(
				caller=self.themmon_button,
				max_height=500,
				items=self.list,
				width_mult=4
			)
		self.menu.open()
	def press(self,args):
		self.menu.dismiss()
		self.themmon(args)
	def nhapdiem(self,*args):
		try:
			diem=float(args[0].text)
			if diem < 0:
				args[0].error=True
				args[0]._anim_current_line_color(args[0].error_color)
				args[0].helper_text="Kí tự này không phải là điểm"
				return
			args[0].helper_text=""
			args[0].error=False
			args[0]._anim_current_line_color(args[0].theme_cls.disabled_hint_text_color)
		except:
			if args[0].text=="":
				args[0].helper_text="Điểm không thể bỏ trống"
			else:
				args[0].helper_text="Kí tự này không phải là điểm"
			args[0].error=True
			args[0]._anim_current_line_color(args[0].error_color)
		self.resize_alldiem(args[0].alldiem)
	def resize_alldiem(self,*args):
		if len(args)>1:
			size=0
			for i in args[1]:
				size+=i.height
			args[0].parent.parent.height=args[0].parent.parent.height-args[0].height+size
			args[0].height=size

		alldiem=[]
		for i in args[0].parent.children:
			if hasattr(i,"ten"):
				if i.ten == "alldiem":
					for j in i.children:
						if type(j) == kivy.uix.gridlayout.GridLayout:
							try:
								alldiem.append([float(j.children[1].text),float(i.cot.heso)])
							except:
								pass
		tongdiem=0
		tongheso=0
		for i in alldiem:
			tongdiem+=i[0]*i[1]
			tongheso+=i[1]
		if tongdiem != 0 and tongheso != 0:
			if len(args[0].parent.parent.children[1].children[0].children)==1:
				args[0].parent.parent.children[1].children[0].add_widget(MDLabel(text=f"ĐTB-Môn:\n{round(tongdiem/tongheso,1)}",halign="center",theme_text_color="Custom",text_color=(102, 255, 102)))
			else:
				args[0].parent.parent.children[1].children[0].children[0].text=f"ĐTB-Môn:\n{round(tongdiem/tongheso,1)}"
		else:
			if len(args[0].parent.parent.children[1].children[0].children)==2:
				args[0].parent.parent.children[1].children[0].remove_widget(args[0].parent.parent.children[1].children[0].children[0])
	def resize_col2(self,*args):
		for i in args[0].children:
			for j in i.children:
				if j.ten == 'adddiem':
					j.width=args[0].width-100
	def resize_diem(self,*args):
		size=0
		for i in args:
			if type(i)==kivy.properties.ObservableList:
				for j in i:
					size+=j.size[1]
			else:
				size+=i.size[1]
		size-=args[0].height
		args[0].parent.height=size+20
		if args[0].parent.height < 300:
			args[0].parent.height=300
	def diem_delete(self,widget_col2):
		widget_col2.parent.parent.remove_widget(widget_col2.parent)
	def update_rect(self,rect,a):
		rect.rect.pos=rect.pos
		rect.rect.size=rect.size
	def load_dulieu_monhoc(self):
		#Hocki1###########################################################################################################################################################
		for mon in self.monhoc:
			self.View=GridLayout(
					rows=1,
					size_hint_y=None,
				)
			self.col1=GridLayout(
				size_hint_x=None,
				width=210,
				cols=1,
			)
			self.View.add_widget(self.col1)
			self.col2=GridLayout(
				cols=1,
			)
			mon.col2hk1=self.col2
			self.col2.bind(size=self.resize_col2)
			self.View.add_widget(self.col2)
			self.View1=GridLayout(
				cols=1,
			)
			monhoc=MDLabel(text=f"Môn Học:\n{mon.ten}",halign="center",theme_text_color="Custom",text_color=(102, 255, 102))
			mon.layout_mon1=monhoc
			self.View1.add_widget(monhoc)
			self.col1.add_widget(self.View1)
			for cot in mon.dhk1:
				cotdiem=MDLabel(text=f"Cột Điểm: {cot.ten}",halign="center",theme_text_color="Custom",text_color=(51/256, 153/256, 102/256))#'Primary', 'Secondary', 'Hint', 'Error', 'Custom', 'ContrastParentBackground'
				cot.layout_cot=cotdiem
				cotdiem.ten="cotdiem"
				cotdiem.cot=cot.ten
				self.View2=GridLayout(
					cols=1,
					row_default_height= 50,
					row_force_default=True,
					size_hint_y=None,
					height=40
				)
				self.View2.add_widget(cotdiem)
				self.col2.add_widget(self.View2)

				all_diem=GridLayout(cols=1,size_hint_y=None,)
				all_diem.cot=cot
				all_diem.ten="alldiem"
				self.col2.add_widget(all_diem)
				all_diem.bind(children=self.resize_alldiem)
			
				layout=Screen(size_hint_y=None,size_hint_x=None,height=50*2)
				add_button=MDIconButton(icon="hospital",pos_hint={'center_x':.5,'center_y':.5})
				add_button.bind(on_press=self.themdiem)
				layout.add_widget(add_button)
				layout.ten="adddiem"
				all_diem.add_widget(layout)
				cot.delete=[all_diem,self.View2]
			self.col2.bind(children=self.resize_diem)
			self.resize_diem(self.col2,self.col2.children)
			for cot in mon.dhk2:
				pass
			with self.View.canvas.before:
				Color(rgba=(0, .4, 0, 0.2))
				self.View.rect=RoundedRectangle(radius=[(40.0, 40.0), (40.0, 40.0), (40.0, 40.0), (40.0, 40.0)])
			self.View.bind(pos=self.update_rect,size=self.update_rect)
			mon.layout_hk1=self.View
		#Hocki2###########################################################################################################################################################
		for mon in self.monhoc:
			self.View=GridLayout(
					rows=1,
					size_hint_y=None,
				)
			self.col1=GridLayout(
				size_hint_x=None,
				width=210,
				cols=1,
			)
			self.View.add_widget(self.col1)
			self.col2=GridLayout(
				cols=1,
			)
			mon.col2hk2=self.col2
			self.col2.bind(size=self.resize_col2)
			self.View.add_widget(self.col2)
			self.View1=GridLayout(
				cols=1,
			)
			monhoc=MDLabel(text=f"Môn Học:\n{mon.ten}",halign="center",theme_text_color="Custom",text_color=(102, 255, 102))
			mon.layout_mon2=monhoc
			self.View1.add_widget(monhoc)
			self.col1.add_widget(self.View1)
			for cot in mon.dhk2:
				cotdiem=MDLabel(text=f"Cột Điểm: {cot.ten}",halign="center",theme_text_color="Custom",text_color=(51/256, 153/256, 102/256))#'Primary', 'Secondary', 'Hint', 'Error', 'Custom', 'ContrastParentBackground'
				cot.layout_cot=cotdiem
				cotdiem.ten="cotdiem"
				cotdiem.cot=cot.ten
				self.View2=GridLayout(
					cols=1,
					row_default_height= 50,
					row_force_default=True,
					size_hint_y=None,
					height=40
				)
				self.View2.add_widget(cotdiem)
				self.col2.add_widget(self.View2)

				all_diem=GridLayout(cols=1,size_hint_y=None,)
				all_diem.cot=cot
				all_diem.ten="alldiem"
				self.col2.add_widget(all_diem)
				all_diem.bind(children=self.resize_alldiem)
		
				layout=Screen(size_hint_y=None,size_hint_x=None,height=50*2)
				add_button=MDIconButton(icon="hospital",pos_hint={'center_x':.5,'center_y':.5})
				add_button.bind(on_press=self.themdiem)
				layout.add_widget(add_button)
				layout.ten="adddiem"
				all_diem.add_widget(layout)
				cot.delete=[all_diem,self.View2]
			self.col2.bind(children=self.resize_diem)
			self.resize_diem(self.col2,self.col2.children)
			with self.View.canvas.before:
				Color(rgba=(0, .4, 0, 0.2))
				self.View.rect=RoundedRectangle(radius=[(40.0, 40.0), (40.0, 40.0), (40.0, 40.0), (40.0, 40.0)])
			self.View.bind(pos=self.update_rect,size=self.update_rect)
			mon.layout_hk2=self.View
	def themmon(self,*args):
		tenmon=args[0]
		#hocki1###################################################################################################################################################
		for mon in self.monhoc:
			if mon.ten == tenmon:
				self.tabdiem.View.add_widget(mon.layout_hk1)
				self.tabdiem.View2.add_widget(mon.layout_hk2)
				diem=[]
				for children in mon.layout_hk1.children[0].children:
					for item in children.children:
						if item.ten=="diem":
							diem.append(item)
				for children in mon.layout_hk2.children[0].children:
					for item in children.children:
						if item.ten=="diem":
							diem.append(item)
				for i in diem:
					i.parent.remove_widget(i)
		try:
			parent=self.themmon_button.parent
			parent.remove_widget(self.themmon_button)
			self.themmon_button=MDIconButton(icon="loupe")
			self.themmon_button.bind(on_press=self.menu_open)
			parent.add_widget(self.themmon_button)
			parent=self.themmon_button2.parent
			parent.remove_widget(self.themmon_button2)
			self.themmon_button2=MDIconButton(icon="loupe")
			self.themmon_button2.bind(on_press=self.menu_open)
			parent.add_widget(self.themmon_button2)
		except:
			self.themmon_button=MDIconButton(icon="loupe")
			self.themmon_button.bind(on_press=self.menu_open)
			self.tabdiem.View.add_widget(self.themmon_button)
			self.themmon_button2=MDIconButton(icon="loupe")
			self.themmon_button2.bind(on_press=self.menu_open)
			self.tabdiem.View2.add_widget(self.themmon_button2)
		###########################################################################################################################################################
	def themdiem(self,*args):
		View3=GridLayout(
			cols=2,
			size_hint_y=None,
			height=100,
			padding=(10, 10),
		)
		diemso=MDTextField(
			font_name="viet",
			text="0",
			halign="center",
			helper_text="",
			helper_text_mode="persistent",
			)
		diemso.bind(text=self.nhapdiem)
		button=MDIconButton(icon="close",size_hint_x=None,width=20)
		button.bind(on_press=self.diem_delete)
		View3.add_widget(diemso)
		View3.add_widget(button)
		View3.ten="diem"
		
		layout=Screen(size_hint_y=None,size_hint_x=None,height=50*2)
		add_button=MDIconButton(icon="hospital",pos_hint={'center_x':.5,'center_y':.5})
		add_button.bind(on_press=self.themdiem)
		layout.add_widget(add_button)
		layout.ten="adddiem"
		try:
			args[0].parent.parent.add_widget(View3)
			args[0].parent.parent.add_widget(layout)
			diemso.alldiem=args[0].parent.parent
			args[0].parent.parent.remove_widget(args[0].parent)
			self.resize_col2(View3.parent.parent)
		except:
			if args[0]==1:hocki=self.tabdiem.View.children[1:]
			else:hocki=self.tabdiem.View2.children[1:]
			for View in hocki:
				if len(View.children[1].children[0].children)==1:
					mon=View.children[1].children[0].children[0].text[9:]
				else:
					mon=View.children[1].children[0].children[1].text[9:]
				if args[1]==mon:
					for View_children in View.children[0].children:
						for item in View_children.children:
							if hasattr(item, 'ten'):
								if type(item) == kivymd.uix.label.MDLabel:
									if args[2]==item.cot:
										diemso.alldiem=button.parent
										diemso.text=str(args[3])
										button.parent.add_widget(View3)
										button.parent.add_widget(layout)
										button.parent.remove_widget(button)
										self.resize_col2(View3.parent.parent)
								if item.ten=="adddiem":
									button=item
		return diemso
	#####################################################################################################################################################




	def touch_up(self,instrance,touch):
		if hasattr(self, 'bubble'):
			self.screen.remove_widget(self.bubble)
	def touch_down(self,instrance,touch):
		pos=(touch.pos[0]-50,touch.pos[1]+20)
		self.touch_pos=pos
	def move_tab(self,tab):
		if hasattr(self, 'bubble'):
			self.screen.remove_widget(self.bubble)
		if tab == "Main":
			self.tab.current=tab
			self.tab.transition.direction="right"
			self.tabdiem.View.clear_widgets()
			self.tabdiem.View2.clear_widgets()
			self.save_tab="main"
		elif tab == "tabdiem":
			self.save_tab="tabdiem"
			self.tab.transition.direction="left"
			self.tab.current="TimeLoad"
	def show_bubble(self):
		if hasattr(self, 'bubble'):
			self.screen.remove_widget(self.bubble)
		self.bubble=Bubble(
				orientation = "horizontal",
				size_hint = (None,None),
				size = (500,120),
				pos=(self.touch_pos[0]-200,self.touch_pos[1])
			)
		button1=BubbleButton(text="Điểm")
		button1.bind(on_press=lambda x: self.move_tab("tabdiem"))
		button2=BubbleButton(text="Xóa học sinh")
		button2.bind(on_press=lambda x: self.xoahocsinh(True,self.data_hocsinh[0],self.data_hocsinh[1],self.data_hocsinh[2],self.data_hocsinh[3]))
		self.bubble.add_widget(button1)
		self.bubble.add_widget(button2)
		self.screen.add_widget(self.bubble)
	def dialog_exit(self,obj):
		obj.parent.parent.parent.parent.dismiss()
	def Load_file(self,instrance):
		self.Filepath=""
		self.open_loadfile(False)
	def select_path(self,path):
		self.Filepath=path
		self.exit_loadfile()
	def select_pathsave(self,path):
		if hasattr(self,"Filepath"):
			self.Filepath_old=self.Filepath
		self.Filepath=path
		if self.Filepath[-3:] != "LTP":
			Filep=setnamefile(self)
			if self.save_tab=="tabdiem":
				self.tabdiem.add_widget(self.Popup_on)
				self.tabdiem.add_widget(Filep)
			else:
				self.screen.add_widget(self.Popup_on)
				self.screen.add_widget(Filep)
			self.Savefile_Layout.close()
			return
		self.save()
	def open_loadfile(self,type):
		if type:
			if platform == 'android':
				self.Savefile_Layout.show("/storage/emulated/0")
			else:
				self.Savefile_Layout.show("/Users/Administrator")
		else:
			if platform == 'android':
				self.Loadfile_Layout.show("/storage/emulated/0")
			else:
				self.Loadfile_Layout.show("/Users/Administrator")
	def exit_loadfile(self,*args):
		self.Loadfile_Layout.close()
		if self.Filepath[-3:] == "LTP":
			self.truong.clear()
			self.monhoc.clear()
			self.loi.clear()
			with open(self.Filepath, 'r', encoding='utf-8') as f:
				for i in f:
					i=i.replace("\n","")
					if ("T::" in i):
						self.themtruong(i[3:])
						tent=i[3:]
					elif ("L::" in i):
						self.themlop(tent,i[3:])
						tenl=i[3:]
					elif ("THS::" in i):
						self.themhs(tent,tenl,i[5:],"0/0/0")
						tenh=i[5:]
					elif ("NTNS::" in i):
						ntns=i[6:]
						for tr in self.truong:
							if tr.ten == tent:
								for lop in tr.lop:
									if lop.ten == tenl:
										for hs in lop.hocsinh:
											if hs.ten == tenh:
												hs.NTNS = ntns
												hs.truong=tr.ten
												hs.lop=lop.ten
					elif ("phai::" in i):
						if i[6::] == "False":
							self.geths(tent,tenl,tenh,ntns).phai=False
					elif ("MHHS" in i):
						self.geths(tent,tenl,tenh,ntns).themmh(i[6:])
					elif ("CDHK1::" in i):
						try:
							index = i.index(",")
							diem=i[index+1:]
						except:
							index=len(i)
							diem=""
						self.geths(tent,tenl,tenh,ntns).diemhs[-1].dhk1.append(Cotdiemhs(i[7:index]))
						self.geths(tent,tenl,tenh,ntns).diemhs[-1].dhk1[-1].diem.extend(diem.split(","))
					elif ("CDHK2" in i):
						try:
							index = i.index(",")
							diem=i[index+1:]
						except:
							index=len(i)
							diem=""
						self.geths(tent,tenl,tenh,ntns).diemhs[-1].dhk2.append(Cotdiemhs(i[7:index]))
						self.geths(tent,tenl,tenh,ntns).diemhs[-1].dhk2[-1].diem.extend(diem.split(","))
					elif ("TenLoi::" in i):
						tenloi=i[8:]
					elif ("Date::" in i):
						self.geths(tent,tenl,tenh,ntns).loi.append(Loihs(tenloi,i[6:]))
					elif ("MH::" in i):
						self.monhoc.append(Monhoc(i[4:]))
					elif ("MHC::" in i):
						if i[5:]=="True":
							self.monhoc[-1].monchinh=True
					elif ("HK1CD::" in i):
						self.monhoc[-1].dhk1.append(Diem(i[7:]))
					elif ("HK1HS::" in i):
						self.monhoc[-1].dhk1[-1].heso=float(i[7:])
					elif ("HK2CD::" in i):
						self.monhoc[-1].dhk2.append(Diem(i[7:]))
					elif ("HK2HS::" in i):
						self.monhoc[-1].dhk2[-1].heso=float(i[7:])
					elif ("TLoi::" in i):
						tloi=i[6:]
					elif ("Diem::" in i):
						self.loi.append(Loi(tloi,float(i[6:])))
			for mon in self.monhoc:
				for cotdiem in mon.dhk1:
					for tr in self.truong:
						for lop in tr.lop:
							for hs in lop.hocsinh:
								for monhs in hs.diemhs:
									if monhs.ten == mon.ten:
										for cotdiemhs in monhs.dhk1:
											if cotdiemhs.ten == cotdiem.ten:
												break
										else:
											monhs.dhk1.append(Cotdiemhs(cotdiem.ten))
				for cotdiem in mon.dhk2:
					for tr in self.truong:
						for lop in tr.lop:
							for hs in lop.hocsinh:
								for monhs in hs.diemhs:
									if monhs.ten == mon.ten:
										for cotdiemhs in monhs.dhk2:
											if cotdiemhs.ten == cotdiem.ten:
												break
										else:
											monhs.dhk2.append(Cotdiemhs(cotdiem.ten))
			self.datatables_rows_update()
			self.load_dulieu_monhoc()
			self.layout_thietlapmon.load_all_cotdiem()
			thongbao("Đã nhập dữ liệu thành công")
		else:
			if self.Filepath!="":
				thongbao("Không thể nhập dữ liệu")
	
	def save(self):
		if not hasattr(self,'Filepath') or self.Filepath=="":
			self.open_loadfile(True)
			return
		with open(self.Filepath, 'w', encoding='utf-8') as f:
			f.write('*************************Thông tin học sinh*************************\n')
			for tr in self.truong:
				f.write('T::'+tr.ten+'\n')
				for lop in tr.lop:
					f.write('L::'+lop.ten+'\n')
					for hs in lop.hocsinh:
						f.write('THS::'+hs.ten+'\n')
						f.write('NTNS::'+hs.NTNS+'\n')
						f.write('phai::'+str(hs.phai)+'\n')
						tmp = 1
						for diemhs in hs.diemhs:
							if tmp == 1:
								f.write('***Điểm***\n')
								tmp = 0
							f.write(f'MHHS::{diemhs.ten}\n')
							for cotdiem in diemhs.dhk1:
								f.write(f'CDHK1::{cotdiem.ten}')
								for diemso in cotdiem.diem:
									f.write(f',{diemso}')
								f.write('\n')
							for cotdiem in diemhs.dhk2:
								f.write(f'CDHK2::{cotdiem.ten}')
								for diemso in cotdiem.diem:
									f.write(f',{diemso}')
								f.write('\n')
						f.write('\n')
						if hs.loi != []:
							f.write('***Lỗi vi phạm***\n')
							for i in hs.loi:
								f.write(f'TenLoi::{i.ten}\n')
								f.write(f'Date::{i.date}\n')
							f.write('\n')
			f.write('*************************Dữ liệu hệ thống môn học*************************\n')
			for mon in self.monhoc:
				f.write('MH::'+mon.ten+'\n')
				f.write(f'MHC::{mon.monchinh}\n')
				for cotdiem in mon.dhk1:
					f.write('HK1CD::'+cotdiem.ten+'\n')
					f.write('HK1HS::'+str(cotdiem.heso)+'\n')
				for cotdiem in mon.dhk2:
					f.write('HK2CD::'+cotdiem.ten+'\n')
					f.write('HK2HS::'+str(cotdiem.heso)+'\n')
			f.write('*************************Dữ liệu định dạng lỗi vi phạm*************************\n')
			for i in self.loi:
				f.write(f'TLoi::{i.ten}\n')
				f.write(f'Diem::{i.diem}\n')
			f.close()
		if self.Savefile_Layout._window_manager != None: self.Savefile_Layout.close()
		thongbao('Đã lưu dữ liệu thành công')
	def xoahocsinh(self,update,tentruong,tenlop,tenhs,NTNS):
		for tr in self.truong:
			if tr.ten == tentruong:
				for lop in tr.lop:
					if lop.ten == tenlop:
						for hs in lop.hocsinh:
							if hs.ten == tenhs and hs.NTNS == NTNS:
								lop.hocsinh.remove(hs)
								if lop.hocsinh==[]:
									tr.lop.remove(lop)
								if tr.lop==[]:
									self.truong.remove(tr)
		if update:self.datatables_rows_update()
	def datatables_rows_update(self,*args):
		listt=[]
		for tr in self.truong:
			for lop in tr.lop:
				for hs in lop.hocsinh:
					row=[]
					row.append(tr.ten)
					row.append(lop.ten)
					row.append(hs.ten)
					row.append(hs.NTNS)
					row.append(hs._phai())
					listt.append(tuple(row))
		self.screen.datatb.row_data=listt
		self.Toolbar.right_action_items=[["account-search",self.lochs]]
	def on_check_press(self,*args):
		if args[1][4]=="Nam":
			args[1][4]=('human-male', 'Nam')
		else:
			args[1][4]=('human-female', 'Nữ')
		index=self.screen.datatb.row_data.index(tuple(args[1]))
		if index in self.screen.datatb.indexx:
			self.screen.datatb.indexx.remove(index)
		else:
			self.screen.datatb.indexx.append(index)
		self.check_press=True
	def selec_row(self,table, row):
		if self.check_press:
			self.check_press=False
		else:
			self.data_hocsinh=[]
			start_index, end_index = row.table.recycle_data[row.index]["range"]
			for i in range(start_index,end_index+1):
				self.data_hocsinh.append(row.table.recycle_data[i]["text"])
			self.show_bubble()
	def themtruong(self,tentruong):
		self.truong.append(Ctruong(tentruong))
	def themlop(self,tentruong,tenlop):
		for o in self.truong:
			if o.ten == tentruong:
				o.lop.append(Clop(tenlop))
				o.tongslop+=1
				return
	def themhs(self,tentruong, tenlop, tenhs, ntns):
		for o in self.truong:
			if o.ten == tentruong:
				for oo in o.lop:
					if oo.ten == tenlop:
						oo.hocsinh.append(Chocsinh(tenhs,ntns))
						oo.tshocsinh+=1
						return
	def geths(self,tentruong,tenlop,tenhs,ntns):
		for tr in self.truong:
			if tr.ten == tentruong:
				for lop in tr.lop:
					if lop.ten == tenlop:
						for hs in lop.hocsinh:
							if hs.ten == tenhs and hs.NTNS == ntns:
								return hs
	def gethesomh(self,hocki,monh,cotd):
	    for mon in self.monhoc:
	        if mon.ten == monh:
	            if hocki == 1:
	                for cotdiem in mon.dhk1:
	                    if cotdiem.ten == cotd:
	                        return float(cotdiem.heso)
	            if hocki == 2:
	                for cotdiem in mon.dhk2:
	                    if cotdiem.ten == cotd:
	                        return float(cotdiem.heso)
	    return False
	def monhocchinh(self):
	    monchinh=[]
	    for mon in self.monhoc:
	        if mon.monchinh==True:
	            monchinh.append(mon.ten)
	    return monchinh
		# pop=Popup(title="a",size_hint=(0.8,0.8),pos_hint={"center_x":0.5,"center_y":0.5},content=layout)
		# pop.open()
	def Load_file_Excel(self,*args):
		if platform == 'android':
			self.Loadfile_Excel_Layout.show("/storage/emulated/0")
		else:
			self.Loadfile_Excel_Layout.show("/Users/Administrator")
	def select_path_excel_loadfile(self,*args):
		self.path_loadfile_excel=args[0]
		self.Loadfile_Excel_Layout.close()
		wb=load_workbook(self.path_loadfile_excel)
		ws_active=wb[wb.get_sheet_names()[0]]
		datahs=[]
		try:
			for hang in range(2,ws_active.max_row+1):
				data=[]
				dont_in_data=False
				for cot in range(1,ws_active.max_column+1):
					for tr in self.truong:
						if ws_active[getcl(cot)+str(hang)].value == tr.ten:
							dont_in_data=True
							break
					else:
						data.append(ws_active[getcl(cot)+str(hang)].value)
					if dont_in_data:break
				if data != []:
					datahs.append(data)
			for hs in datahs:
				self.truong.append(Ctruong(hs[0]))
				tmp=False
				for tr in self.truong:
					for lop in tr.lop:
						if lop.ten.lower() == hs[3].lower():
							self.truong[-1].lop.append(Clop(lop.ten))
							tmp=True
							break
					if tmp:
						break
				else:
					self.truong[-1].lop.append(Clop(hs[3]))
				self.truong[-1].lop[-1].hocsinh.append(Chocsinh(hs[1],hs[2]))
				if hs[4]=="Nữ":
					self.truong[-1].lop[-1].hocsinh[-1].phai=False
			self.datatables_rows_update()
		except:
			thongbao('Lỗi không thể đọc tài liệu ở danh sách học sinh')
			return



		alldata=[]
		try:
			for mon_sheet in wb.get_sheet_names()[1:]:
				sheet=wb[mon_sheet]
				skip=False
				all_data_diem=[]
				alldata.append(sheet.title)
				hs=""
				if type(sheet[getcl(2)+"1"].value)!=bool:
					thongbao(f"Lỗi ở trang tính {sheet.title}, không thể định dạng môn học chính",duration=5)
					return
				alldata.append(sheet[getcl(2)+"1"].value)
				for cot in range(1,sheet.max_column+1):
					if sheet[getcl(cot)+"2"].value != None:
						if  sheet[getcl(cot)+"2"].value.lower() == "học kì i":
							cotHK1=cot
						if sheet[getcl(cot)+"2"].value.lower() == "học kì ii":
							cotHK2=cot
				for hang in range(5,sheet.max_row+1):
					for cot in range(1,sheet.max_column+1):
						if cot==1 and sheet[getcl(cot)+str(hang)].value!=None:
							for tr in self.truong:
								if tr.ten.lower() == sheet[getcl(cot)+str(hang)].value.lower():
									if hs!="":
										all_data_diem.append(hs)
									hs=[sheet[getcl(cot)+str(hang)].value]
									skip=False
									break
							else:
								skip=True
						if skip:
							break
						if cot > 2 and sheet[getcl(cot)+str(hang)].value!=None:
							if cot < cotHK2:
								hs.append(["Học kì I",sheet[getcl(cot)+"3"].value,sheet[getcl(cot)+"4"].value,sheet[getcl(cot)+str(hang)].value])
							else:
								hs.append(["Học kì II",sheet[getcl(cot)+"3"].value,sheet[getcl(cot)+"4"].value,sheet[getcl(cot)+str(hang)].value])
				if hs!="":
					all_data_diem.append(hs)
				alldata.append(all_data_diem)
		except:
			thongbao(f'Lỗi không thể đọc tài liệu từ trang tính {sheet.title}',duration=5)
			return
		for data in alldata:
			if type(data)==str:
				for mon in self.monhoc:
					if mon.ten.lower()==data.lower():
						mon.monchinh=alldata[alldata.index(data)+1]
						for item in alldata[alldata.index(data)+2]:
							for i in item[1:]:
								if i[0] == "Học kì I":
									for cot in mon.dhk1:
										if cot.ten.lower() == i[1].lower():
											break
									else:
										#tao cot diem
										mon.dhk1.append(Diem(i[1],i[2]))
								if i[0] == "Học kì II":
									for cot in mon.dhk2:
										if cot.ten.lower() == i[1].lower():
											break
									else:
										#tao cot diem
										mon.dhk2.append(Diem(i[1],i[2]))
						break
				else:
					#tao mon hoc
					self.monhoc.append(Monhoc(data))
					self.monhoc[-1].monchinh=alldata[alldata.index(data)+1]
					for item in alldata[alldata.index(data)+2]:
						for i in item[1:]:
							if i[0] == "Học kì I":
								for cot in self.monhoc[-1].dhk1:
									if cot.ten.lower() == i[1].lower():
										break
								else:
									#tao cot diem
									self.monhoc[-1].dhk1.append(Diem(i[1],i[2]))
							if i[0] == "Học kì II":
								for cot in self.monhoc[-1].dhk2:
									if cot.ten.lower() == i[1].lower():
										break
								else:
									#tao cot diem
									self.monhoc[-1].dhk2.append(Diem(i[1],i[2]))

		for data in alldata:
			if type(data)==str:
				for item in alldata[alldata.index(data)+2]:
					hs=self.geths_mahs(item[0])
					for mon in hs.diemhs:
						if mon.ten.lower()==data.lower():

							for cotdiem in mon.dhk1:
								cotdiem.diem=[]
							for cotdiem in mon.dhk2:#remove all diem
								cotdiem.diem=[]

							for i in item[1:]:
								if i[0]=='Học kì I':
									for cotdiem in mon.dhk1:
										if cotdiem.ten.lower()==i[1].lower():
											break
									else:
										mon.dhk1.append(Cotdiemhs(i[1]))
									for cotdiem in mon.dhk1:
										if cotdiem.ten.lower() == i[1].lower():
											cotdiem.diem.append(float(i[3]))
								if i[0]=='Học kì II':
									for cotdiem in mon.dhk2:
										if cotdiem.ten.lower()==i[1].lower():
											break
									else:
										mon.dhk2.append(Cotdiemhs(i[1]))
									for cotdiem in mon.dhk2:
										if cotdiem.ten.lower() == i[1].lower():
											cotdiem.diem.append(float(i[3]))
							break
					else:
						#them mon hoc cho hs
						for mon in self.monhoc:
							if mon.ten.lower() == data.lower():
								hs.diemhs.append(Diemhs(mon.ten))
								for i in item[1:]:
									if i[0]=='Học kì I':
										for cotdiem in hs.diemhs[-1].dhk1:
											if cotdiem.ten.lower()==i[1].lower():
												break
										else:
											hs.diemhs[-1].dhk1.append(Cotdiemhs(i[1]))
										for cotdiem in hs.diemhs[-1].dhk1:
											if cotdiem.ten.lower() == i[1].lower():
												cotdiem.diem.append(float(i[3]))
									if i[0]=='Học kì II':
										for cotdiem in hs.diemhs[-1].dhk2:
											if cotdiem.ten.lower()==i[1].lower():
												break
										else:
											hs.diemhs[-1].dhk2.append(Cotdiemhs(i[1]))
										for cotdiem in hs.diemhs[-1].dhk2:
											if cotdiem.ten.lower() == i[1].lower():
												cotdiem.diem.append(float(i[3]))
		self.load_dulieu_monhoc()
		self.layout_thietlapmon.load_all_cotdiem()
		thongbao("Đã nhập dữ liệu thành công")

	def geths_mahs(self,Mahs):
		for tr in self.truong:
			for lop in tr.lop:
				for hs in lop.hocsinh:
					if tr.ten.lower()==Mahs.lower():
						return hs
	def tongsohs(self):
		cout=0
		for tr in self.truong:
			for lop in tr.lop:
				for hs in lop.hocsinh:
					cout+=1
		return cout
	def tong_hocsinh_gioi(self,hocki,lopp):
		cout=0
		for tr in self.truong:
			for lop in tr.lop:
				if lop.ten == lopp or lopp == 'Toàn khối':
					for hs in lop.hocsinh:
						if hs.xeploai(hocki) == "Giỏi":
							cout+=1
		return cout
	def tong_hocsinh_kha(self,hocki,lopp):
		cout=0
		for tr in self.truong:
			for lop in tr.lop:
				if lop.ten == lopp or lopp == 'Toàn khối':
					for hs in lop.hocsinh:
						if hs.xeploai(hocki) == "Khá":
							cout+=1
		return cout
	def tong_hocsinh_trungbinh(self,hocki,lopp):
		cout=0
		for tr in self.truong:
			for lop in tr.lop:
				if lop.ten == lopp or lopp == 'Toàn khối':
					for hs in lop.hocsinh:
						if hs.xeploai(hocki) == "Trung bình":
							cout+=1
		return cout
	def tong_hocsinh_yeu(self,hocki,lopp):
		cout=0
		for tr in self.truong:
			for lop in tr.lop:
				if lop.ten == lopp or lopp == 'Toàn khối':
					for hs in lop.hocsinh:
						if hs.xeploai(hocki) == "Yếu":
							cout+=1
		return cout
	def tong_hocsinh_kem(self,hocki,lopp):
		cout=0
		for tr in self.truong:
			for lop in tr.lop:
				if lop.ten == lopp or lopp == 'Toàn khối':
					for hs in lop.hocsinh:
						if hs.xeploai(hocki) == "Kém":
							cout+=1
		return cout
class getfilepath_Excel(GridLayout):
	def __init__(self,*args,**kwargn):
		super(getfilepath_Excel,self).__init__(**kwargn)
		Window.bind(size=self.resize)
		self.width=Window.width*70/100
		self.rows=1
		self.pos_hint={"center_x":.5,"center_y":.5}
		self.size_hint=(None,None)
		self.height=200
		self.padding=20
		self.textfield=MDTextField(
				font_name="viet",
				hint_text="Tên file dữ liệu",
				pos_hint={'center_x':.5,'center_y':.5},
				font_size=sp(15),
				text_color=(74/256, 173/256, 54/256, 1),
				halign="center",
			)
		self.add_widget(self.textfield)
		with self.canvas.before:
			Color(rgba=(62/256, 112/256, 128/256, 1))
			self.rect=RoundedRectangle(radius=[(40.0, 40.0), (40.0, 40.0), (40.0, 40.0), (40.0, 40.0)])
		self.bind(pos=app.update_rect,size=app.update_rect)
		self.xong=MDFillRoundFlatButton(text="Lưu")
		self.xong.bind(on_press=self.xong_)
		self.huy=MDFillRoundFlatButton(text="Hủy")
		self.huy.bind(on_press=self.huy_)
		self.add_widget(MDLabel(text=".xlsx",size_hint_x=None,size_hint_y=None,height=100,width=100,halign="auto"))
		self.add_widget(self.huy)
		self.add_widget(self.xong)
	def huy_(self,*args):
		self.parent.remove_widget(app.Popup_on)
		self.parent.remove_widget(self)
	def xong_(self,*args):
		self.parent.remove_widget(app.Popup_on)
		self.parent.remove_widget(self)
		if self.textfield.text=='':
			thongbao('Chưa nhập tên file')
			return
		app.screen.xuatdulieu(app.screen._path_+f"/{self.textfield.text}",app.screen.data_xuatfile_excel)
	def resize(self,*args):
		self.width=Window.width*70/100
class setnamefile(GridLayout):
	def __init__(self,*args,**kwargn):
		super(setnamefile,self).__init__(**kwargn)
		self.app=args[0]
		Window.bind(size=self.resize)
		self.width=Window.width*70/100
		self.rows=1
		self.pos_hint={"center_x":.5,"center_y":.5}
		self.size_hint=(None,None)
		self.height=200
		self.padding=20
		self.textfield=MDTextField(
				font_name="viet",
				hint_text="Tên file mới",
				pos_hint={'center_x':.5,'center_y':.5},
				font_size=sp(15),
				text_color=(74/256, 173/256, 54/256, 1),
				halign="center",
			)
		self.add_widget(self.textfield)
		with self.canvas.before:
			Color(rgba=(62/256, 112/256, 128/256, 1))
			self.rect=RoundedRectangle(radius=[(40.0, 40.0), (40.0, 40.0), (40.0, 40.0), (40.0, 40.0)])
		self.bind(pos=self.app.update_rect,size=self.app.update_rect)
		self.xong=MDFillRoundFlatButton(text="Lưu")
		self.xong.bind(on_press=self.xong_)
		self.huy=MDFillRoundFlatButton(text="Hủy")
		self.huy.bind(on_press=self.huy_)
		self.add_widget(MDLabel(text=".LTP",size_hint_x=None,size_hint_y=None,height=100,width=100,halign="auto"))
		self.add_widget(self.huy)
		self.add_widget(self.xong)
	def huy_(self,*args):
		self.parent.remove_widget(self.app.Popup_on)
		self.parent.remove_widget(self)
		if hasattr(self.app,"Filepath_old"):
			self.app.Filepath=self.app.Filepath_old
		else:
			self.app.Filepath=""
	def xong_(self,*args):
		self.parent.remove_widget(self.app.Popup_on)
		self.parent.remove_widget(self)
		if self.textfield.text == "":
			self.app.Filepath+="/QLSV.LTP"
		else:
			self.app.Filepath+=f"/{self.textfield.text}.LTP"
		self.app.save()
	def resize(self,*args):
		self.width=Window.width*70/100
class layout_timhocsinh_MHS(GridLayout):
	def __init__(self,*args,**kwargn):
		super(layout_timhocsinh_MHS,self).__init__(**kwargn)
		self.rows=1
		self.size_hint=(None,None)
		self.pos_hint={"center_x":.5,"center_y":.5}
		with self.canvas.before:
			Color(rgba=(62/256, 112/256, 128/256, 1))
			self.rect=RoundedRectangle(radius=[(40.0, 40.0), (40.0, 40.0), (40.0, 40.0), (40.0, 40.0)])
		self.bind(pos=app.update_rect,size=app.update_rect)
		self.textfield=MDTextField(
				text="",
				hint_text="Mã học sinh cần tìm",
				size_hint_x=None,
				width=Window.width*70/100,
				halign="center",
				text_color=(74/256, 173/256, 54/256, 1),
			)
		self.textfield.bind(focus=self.on_edit,text=self.on_edit)
		self.add_widget(self.textfield)
		button=MDFillRoundFlatButton(text="tìm")
		button.bind(on_press=self.tim)
		self.add_widget(button)
		w=0
		h=0
		for i in self.children:
			w+=i.width
			if h<i.height:
				h=i.height+120
		self.width=w
		self.height=h

		self.menu=MDDropdownMenu(
					caller=self.textfield,
					max_height=500,
					width_mult=4,
					position="bottom"
				)
	def menu_on(self,t):
		self.textfield.text=t
		self.menu.dismiss()
	def on_edit(self,*args):
		if self.textfield.focus:self.menu.dismiss()
		if self.textfield.focus and args[1] != "":
			self.menu.items=[]
			for tr in app.truong:
				if self.textfield.text.lower() in tr.ten.lower():
					self.menu.items.append(
							{
								"text":tr.ten,
								"viewclass": "OneLineListItem",
								"on_release":lambda x=tr.ten:self.menu_on(x),
							},
						)
			if self.menu.items!=[]:self.menu.open()
		elif self.textfield.focus:
			self.menu.items=[]
			for tr in app.truong:
				self.menu.items.append(
						{
							"text":tr.ten,
							"viewclass": "OneLineListItem",
							"on_release":lambda x=tr.ten:self.menu_on(x),
						},
					)
			if self.menu.items!=[]:self.menu.open()
	def tim(self,*args):
		app.screen.remove_widget(self)
		app.screen.remove_widget(app.Popup_on)
		listt=[]
		for tr in app.truong:
			if tr.ten.lower() == self.textfield.text.lower():
				for lop in tr.lop:
					for hs in lop.hocsinh:
						row=[]
						row.append(tr.ten)
						row.append(lop.ten)
						row.append(hs.ten)
						row.append(hs.NTNS)
						row.append(hs._phai())
						listt.append(tuple(row))
		if listt == []:
			thongbao("Mã học sinh không tồn tại")
			return
		app.screen.datatb.row_data=listt
		app.Toolbar.right_action_items=[["close",app.datatables_rows_update]]
class layout_timhocsinh_ten(GridLayout):
	def __init__(self,*args,**kwargn):
		super(layout_timhocsinh_ten,self).__init__(**kwargn)
		self.rows=1
		self.size_hint=(None,None)
		self.pos_hint={"center_x":.5,"center_y":.5}
		with self.canvas.before:
			Color(rgba=(62/256, 112/256, 128/256, 1))
			self.rect=RoundedRectangle(radius=[(40.0, 40.0), (40.0, 40.0), (40.0, 40.0), (40.0, 40.0)])
		self.bind(pos=app.update_rect,size=app.update_rect)
		self.textfield=MDTextField(
				text="",
				hint_text="Tên học sinh cần tìm",
				size_hint_x=None,
				width=Window.width*70/100,
				halign="center",
				text_color=(74/256, 173/256, 54/256, 1),
			)
		self.textfield.bind(focus=self.on_edit,text=self.on_edit)
		self.add_widget(self.textfield)
		button=MDFillRoundFlatButton(text="tìm")
		button.bind(on_press=self.tim)
		self.add_widget(button)
		w=0
		h=0
		for i in self.children:
			w+=i.width
			if h<i.height:
				h=i.height+120
		self.width=w
		self.height=h

		self.menu=MDDropdownMenu(
					caller=self.textfield,
					max_height=500,
					width_mult=4,
					position="bottom"
				)
	def menu_on(self,t):
		self.textfield.text=t
		self.menu.dismiss()
	def on_edit(self,*args):
		if self.textfield.focus:self.menu.dismiss()
		if self.textfield.focus and args[1] != "":
			self.menu.items=[]
			for tr in app.truong:
				for lop in tr.lop:
					for hs in lop.hocsinh:
						if self.textfield.text.lower() in hs.ten.lower():
							self.menu.items.append(
									{
										"text":hs.ten,
										"viewclass": "OneLineListItem",
										"on_release":lambda x=hs.ten:self.menu_on(x),
									},
								)
			if self.menu.items!=[]:self.menu.open()
		elif self.textfield.focus:
			self.menu.items=[]
			for tr in app.truong:
				for lop in tr.lop:
					for hs in lop.hocsinh:
						self.menu.items.append(
								{
									"text":hs.ten,
									"viewclass": "OneLineListItem",
									"on_release":lambda x=hs.ten:self.menu_on(x),
								},
							)
			if self.menu.items!=[]:self.menu.open()
	def tim(self,*args):
		app.screen.remove_widget(self)
		app.screen.remove_widget(app.Popup_on)
		listt=[]
		for tr in app.truong:
			for lop in tr.lop:
				for hs in lop.hocsinh:
					if hs.ten.lower() == self.textfield.text.lower():
						row=[]
						row.append(tr.ten)
						row.append(lop.ten)
						row.append(hs.ten)
						row.append(hs.NTNS)
						row.append(hs._phai())
						listt.append(tuple(row))
		if listt == []:
			thongbao("Tên học sinh không tồn tại")
			return
		app.screen.datatb.row_data=listt
		app.Toolbar.right_action_items=[["close",app.datatables_rows_update]]
class layout_lochs(GridLayout):
	def __init__(self,*args,**kwargn):
		super(layout_lochs,self).__init__(**kwargn)
		self.rows=1
		self.size_hint=(None,None)
		self.pos_hint={"center_x":.5,"center_y":.5}
		with self.canvas.before:
			Color(rgba=(62/256, 112/256, 128/256, 1))
			self.rect=RoundedRectangle(radius=[(40.0, 40.0), (40.0, 40.0), (40.0, 40.0), (40.0, 40.0)])
		self.bind(pos=app.update_rect,size=app.update_rect)
		
		self.mon_button=MDFillRoundFlatButton(text="chọn môn",pos_hint={"top":1})
		self.mon_button.bind(on_press=self.menu_mon)
		self.mon_hk=Screen(size_hint=(.1,None))
		self.hocki=MDFillRoundFlatButton(text="Học kì I",pos_hint={"top":0})
		self.hocki.bind(on_release=self.hocki_)
		self.mon_hk.add_widget(self.mon_button)
		self.mon_hk.add_widget(self.hocki)
		self.add_widget(self.mon_hk)

		self.min=MDTextField(
				text="",
				hint_text="Min",
				helper_text="",
				helper_text_mode="persistent",
				width=60,
				pos_hint={'top':0,'center_x':.2},
				size_hint_x=None,
				halign="center",
				text_color=(74/256, 173/256, 54/256, 1),
			)
		self.min.bind(focus=self.check_error,text=self.check_error)
		self.max=MDTextField(
				text="",
				hint_text="Max",
				helper_text="",
				pos_hint={'top':0,'center_x':.8},
				size_hint_x=None,
				helper_text_mode="persistent",
				width=60,
				halign="center",
				text_color=(74/256, 173/256, 54/256, 1),
			)
		self.max.bind(focus=self.check_error,text=self.check_error)
		tmp=Screen(size_hint=(None,None),width=Window.width*30/100)
		tmp.add_widget(self.min)
		tmp.add_widget(MDLabel(
				text="-",
				pos_hint={'top':0,'center_x':.5},
				halign="center",
				size_hint_x=None,width=70,
				size_hint_y=None,height=100,
			))
		tmp.add_widget(self.max)
		self.add_widget(tmp)
		
		button=MDFillRoundFlatButton(text="Lọc",pos_hint={"center_x":.5,"top":.0})
		button.bind(on_release=self.loc)
		tmp=Screen(size_hint=(.1,None))
		tmp.add_widget(button)
		self.add_widget(tmp)
		w=0
		h=0
		for i in self.children:
			w+=i.width+50
			if h<i.height:
				h=i.height+120
		self.width=w
		self.height=h
		self.menu=MDDropdownMenu(
					caller=self.mon_hk,
					max_height=500,
					width_mult=4,
					position="bottom"
				)
	def check_error(self,*args):
		tf=args[0]
		if tf.focus and tf.text!="":
			try:
				diem=float(tf.text)
				if diem < 0:
					tf.error=True
					tf.helper_text="Chỉ nhập số nguyên dương"
					tf._anim_current_line_color(tf.error_color)
					return
				tf.error=False
				tf.helper_text=""
				tf._anim_current_line_color(tf.theme_cls.disabled_hint_text_color)
			except:
				tf.error=True
				tf.helper_text="Chỉ nhập điểm số"
				tf._anim_current_line_color(tf.error_color)
		elif tf.focus or tf.text=="":
			tf.error=False
			tf.helper_text=""
			tf._anim_current_line_color(tf.theme_cls.disabled_hint_text_color)
		else:
			tf.helper_text=""
	def hocki_(self,*args):
		if args[0].text == "Học kì I":
			args[0].text = "Học kì II"
		else:
			args[0].text = "Học kì I"
	def menu_mon(self,*args):
		self.menu.items=[]
		for mon in app.monhoc:
			self.menu.items.append(
					{
						"text":mon.ten,
						"viewclass": "OneLineListItem",
						"on_release":lambda x=mon.ten:self.slec_mon(x),
					}
				)
		if self.menu.items!=[]:self.menu.open()
		else:
			thongbao("Không có dữ liệu môn học")
	def slec_mon(self,*args):
		self.mon_button.text=args[0]
		self.menu.dismiss()
	def loc(self,*args):
		app.screen.remove_widget(self)
		app.screen.remove_widget(app.Popup_on)
		if self.min.error or self.max.error:
			thongbao('Lỗi điểm cần lọc')
			return
		if self.mon_button.text == "chọn môn":
			thongbao("Chưa chọn môn học")
			return
		if self.min.text == "" and self.max.text == "":
			thongbao("Chưa nhập điểm cần lọc")
			return
		data=[]
		for tr in app.truong:
			for lop in tr.lop:
				for hs in lop.hocsinh:
					for mon in hs.diemhs:
						if mon.ten == self.mon_button.text:
							if self.hocki.text == "Học kì I":
								if self.min.text != "" and self.max.text != "":
									if mon.dtbmon(1) >= float(self.min.text) and mon.dtbmon(1) <= float(self.max.text):
										data.append((tr.ten,lop.ten,hs.ten,hs.NTNS,hs._phai()))
								elif self.min.text != "":
									if mon.dtbmon(1) >= float(self.min.text):
										data.append((tr.ten,lop.ten,hs.ten,hs.NTNS,hs._phai()))
								else:
									if mon.dtbmon(1) <= float(self.max.text):
										data.append((tr.ten,lop.ten,hs.ten,hs.NTNS,hs._phai()))
							else:
								if self.min.text != "" and self.max.text != "":
									if mon.dtbmon(2) >= float(self.min.text) and mon.dtbmon(2) <= float(self.max.text):
										data.append((tr.ten,lop.ten,hs.ten,hs.NTNS,hs._phai()))
								elif self.min.text != "":
									if mon.dtbmon(2) >= float(self.min.text):
										data.append((tr.ten,lop.ten,hs.ten,hs.NTNS,hs._phai()))
								else:
									if mon.dtbmon(2) <= float(self.max.text):
										data.append((tr.ten,lop.ten,hs.ten,hs.NTNS,hs._phai()))
		if data == []:
			thongbao("Không có học sinh nào đạt số điểm này")
			return
		app.screen.datatb.row_data=data
		app.Toolbar.right_action_items=[["close",app.datatables_rows_update]]
if __name__ == "__main__":
	app=Myapp()
	app.run()